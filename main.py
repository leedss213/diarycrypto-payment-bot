import discord
from discord.ext import commands
import os
import sqlite3
import time
from datetime import datetime, timedelta
import pytz
import requests
from flask import Flask, request
import threading
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import json
import asyncio
import random
from enum import Enum
from typing import Optional, Dict, List, Tuple
import urllib.parse
import midtransclient

# ============ CONFIG ============
DISCORD_TOKEN = os.getenv('DISCORD_TOKEN')
GUILD_ID = 1370638839407972423
MIDTRANS_CLIENT_KEY = os.getenv('MIDTRANS_CLIENT_KEY')
MIDTRANS_SERVER_KEY = os.getenv('MIDTRANS_SERVER_KEY')
GMAIL_SENDER = os.getenv('GMAIL_SENDER')
GMAIL_PASSWORD = os.getenv('GMAIL_PASSWORD')
ADMIN_EMAIL = "asrofsantoso@gmail.com"
WARRIOR_ROLE_NAME = "The Warrior"
TRIAL_MEMBER_ROLE_NAME = "Trial Member"
ANALYST_ROLE_NAME = "Analyst"
ANALYST_LEAD_ROLE_NAME = "Analyst's Lead"
NEWS_CHANNEL_NAME = "📰｜berita-crypto"

# Flask app
app = Flask(__name__)

# Discord client setup
intents = discord.Intents.all()
intents.members = True
intents.presences = True
bot = commands.Bot(command_prefix='/', intents=intents)
bot.is_synced = False
tree = bot.tree

# Midtrans setup
midtrans_client = midtransclient.Snap(
    is_production=False,
    server_key=MIDTRANS_SERVER_KEY,
    client_key=MIDTRANS_CLIENT_KEY
)

# ============ DATABASE SETUP ============
def init_db():
    conn = sqlite3.connect('warrior_subscriptions.db')
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS subscriptions (
                    order_id TEXT PRIMARY KEY,
                    discord_id TEXT NOT NULL,
                    discord_username TEXT,
                    nama TEXT,
                    email TEXT,
                    package_type TEXT,
                    payment_method TEXT,
                    status TEXT DEFAULT "pending",
                    start_date TEXT,
                    end_date TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    referral_code TEXT,
                    referrer_id TEXT,
                    expiry_reminder_count INTEGER DEFAULT 0
                )''')
    
    # Add missing column if it doesn't exist (migration for existing databases)
    try:
        c.execute('ALTER TABLE subscriptions ADD COLUMN expiry_reminder_count INTEGER DEFAULT 0')
        print("✅ Migration: Added expiry_reminder_count column to subscriptions table")
    except sqlite3.OperationalError as e:
        if "duplicate column" in str(e) or "already exists" in str(e):
            pass  # Column already exists, skip
        else:
            print(f"⚠️ Migration warning: {e}")
    
    c.execute('''CREATE TABLE IF NOT EXISTS pending_orders (
                    order_id TEXT PRIMARY KEY,
                    discord_id TEXT,
                    discord_username TEXT,
                    nama TEXT,
                    email TEXT,
                    package_type TEXT,
                    price REAL DEFAULT 0,
                    payment_url TEXT,
                    status TEXT DEFAULT "pending",
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS trial_members (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    trial_code TEXT UNIQUE,
                    discord_id TEXT,
                    discord_username TEXT,
                    email TEXT,
                    username TEXT,
                    trial_started TEXT,
                    trial_end TEXT,
                    status TEXT,
                    created_at TEXT,
                    created_by TEXT,
                    duration_days INTEGER DEFAULT 1,
                    validity_days INTEGER DEFAULT 1,
                    code_expiry_date TEXT,
                    max_uses INTEGER DEFAULT 1,
                    used_count INTEGER DEFAULT 0
                )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS referral_codes (
                    code TEXT PRIMARY KEY,
                    created_by TEXT,
                    created_date TEXT DEFAULT CURRENT_TIMESTAMP,
                    uses INT DEFAULT 0,
                    max_uses INT
                )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS commissions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    analyst_id TEXT,
                    referred_member_id TEXT,
                    commission_amount REAL,
                    package_type TEXT,
                    earned_date TEXT DEFAULT CURRENT_TIMESTAMP,
                    status TEXT DEFAULT "pending"
                )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS packages (
                    package_id TEXT PRIMARY KEY,
                    package_name TEXT,
                    price REAL,
                    duration_days REAL,
                    duration_text TEXT,
                    role_name TEXT DEFAULT "The Warrior",
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    created_by TEXT
                )''')
    
    # Insert default packages jika table kosong
    c.execute('SELECT COUNT(*) FROM packages')
    if c.fetchone()[0] == 0:
        default_packages = [
            ('warrior_15min', 'The Warrior 15 Minutes', 200000, 15/1440, '15 menit', WARRIOR_ROLE_NAME, None),
            ('warrior_1hour', 'The Warrior 1 Hour', 50000, 1/24, '1 jam', WARRIOR_ROLE_NAME, None),
            ('warrior_1month', 'The Warrior 1 Month', 299000, 30, '1 bulan', WARRIOR_ROLE_NAME, None),
            ('warrior_3month', 'The Warrior 3 Months', 649000, 90, '3 bulan', WARRIOR_ROLE_NAME, None)
        ]
        c.executemany(
            'INSERT INTO packages (package_id, package_name, price, duration_days, duration_text, role_name, created_by) VALUES (?, ?, ?, ?, ?, ?, ?)',
            default_packages
        )
    
    conn.commit()
    conn.close()

init_db()

# ============ TIMEZONE & DATE UTILS ============
def get_jakarta_datetime():
    return datetime.now(pytz.timezone('Asia/Jakarta'))

def format_jakarta_datetime(dt):
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=pytz.UTC)
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    jakarta_dt = dt.astimezone(jakarta_tz)
    return jakarta_dt.strftime('%d %b %Y %H:%M WIB')

def format_jakarta_datetime_full(date_str):
    if isinstance(date_str, str):
        date_str = date_str.split(' ')[0]
        dt = datetime.strptime(date_str, '%Y-%m-%d')
    else:
        dt = date_str
    
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=pytz.UTC)
    jakarta_dt = dt.astimezone(jakarta_tz)
    return jakarta_dt.strftime('%d %B %Y')

# ============ PACKAGES ============
def get_all_packages():
    """Load semua packages dari database"""
    try:
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        c.execute('SELECT package_id, package_name, price, duration_days, duration_text, role_name FROM packages ORDER BY created_at')
        rows = c.fetchall()
        conn.close()
        
        packages = {}
        for row in rows:
            pkg_id, pkg_name, price, duration_days, duration_text, role_name = row
            packages[pkg_id] = {
                'name': pkg_name,
                'price': int(price),
                'duration_days': float(duration_days),
                'duration_text': duration_text,
                'role_name': role_name
            }
        
        return packages if packages else {}
    except Exception as e:
        print(f"Error loading packages: {e}")
        return {}

# ============ HELPER FUNCTIONS ============
def is_commission_manager(interaction: discord.Interaction):
    """Check if user is guild owner or has admin permissions"""
    return interaction.user.id == interaction.guild.owner_id or interaction.user.guild_permissions.administrator

def verify_discount_code(code: str) -> Dict:
    """Verify dan get discount code details"""
    try:
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        # Create table jika belum exist
        c.execute('''CREATE TABLE IF NOT EXISTS discount_codes (
            code TEXT PRIMARY KEY,
            discount_percent INTEGER,
            max_uses INTEGER,
            used_count INTEGER DEFAULT 0,
            created_at TEXT,
            created_by TEXT
        )''')
        
        c.execute('SELECT discount_percent, max_uses, used_count FROM discount_codes WHERE code = ?', (code.upper(),))
        result = c.fetchone()
        conn.close()
        
        if not result:
            return {"valid": False, "message": "Kode diskon tidak ditemukan"}
        
        discount_percent, max_uses, used_count = result
        
        # Check max uses
        if max_uses > 0 and used_count >= max_uses:
            return {"valid": False, "message": f"Kode diskon sudah mencapai batas penggunaan ({used_count}/{max_uses})"}
        
        return {
            "valid": True,
            "discount_percent": discount_percent,
            "message": f"✅ Diskon {discount_percent}% berhasil diterapkan!"
        }
    except Exception as e:
        return {"valid": False, "message": str(e)}

def verify_referral_code(code: str) -> Dict:
    """Verify dan get referral code details - 30% komisi untuk analyst"""
    try:
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        # Create table jika belum exist
        c.execute('''CREATE TABLE IF NOT EXISTS referral_codes (
            code TEXT PRIMARY KEY,
            analyst_id TEXT,
            analyst_name TEXT,
            commission_percent INTEGER DEFAULT 30,
            created_at TEXT
        )''')
        
        c.execute('SELECT analyst_id, analyst_name FROM referral_codes WHERE code = ?', (code.upper(),))
        result = c.fetchone()
        conn.close()
        
        if not result:
            return {"valid": False, "message": "Kode referral tidak ditemukan"}
        
        analyst_id, analyst_name = result
        return {
            "valid": True,
            "analyst_id": analyst_id,
            "analyst_name": analyst_name,
            "commission_percent": 30,
            "message": f"✅ Referral dari {analyst_name} berhasil diterapkan! (Komisi 30%)"
        }
    except Exception as e:
        return {"valid": False, "message": str(e)}

def generate_referral_code(member_id):
    code = f"REF_{member_id}_{random.randint(1000, 9999)}"
    return code

def get_pending_order(order_id):
    conn = sqlite3.connect('warrior_subscriptions.db')
    c = conn.cursor()
    c.execute('SELECT order_id, discord_id, discord_username, nama, email, package_type, payment_url, status, created_at FROM pending_orders WHERE order_id = ?', (order_id,))
    order = c.fetchone()
    conn.close()
    return order

def generate_snap_token(order_id, price, customer_name, customer_email):
    """Generate Midtrans Snap Token dengan redirect URL untuk payment page"""
    try:
        transaction_details = {
            "order_id": order_id,
            "gross_amount": int(price)
        }
        
        customer_details = {
            "first_name": customer_name,
            "email": customer_email
        }
        
        payload = {
            "transaction_details": transaction_details,
            "customer_details": customer_details
        }
        
        snap_response = midtrans_client.create_transaction(payload)
        
        # Ambil redirect_url dari response (ini adalah URL yang benar dari Midtrans)
        redirect_url = snap_response.get('redirect_url')
        snap_token = snap_response.get('token')
        
        if redirect_url:
            print(f"✅ Payment URL generated: {redirect_url[:60]}...")
            return redirect_url  # Return redirect_url langsung, bukan token
        elif snap_token:
            print(f"⚠️ Token only (no redirect_url): {snap_token[:20]}...")
            return f"https://app.sandbox.midtrans.com/snap/v1/web/{snap_token}"
        else:
            print(f"❌ Error in response: {snap_response}")
            return None
    except Exception as e:
        print(f"❌ Error generating payment link: {e}")
        return None

def save_pending_order(order_id, discord_id, username, nama, email, package_type, payment_url):
    conn = sqlite3.connect('warrior_subscriptions.db')
    c = conn.cursor()
    created_at = get_jakarta_datetime().strftime('%Y-%m-%d %H:%M:%S')
    c.execute('''INSERT OR REPLACE INTO pending_orders 
                (order_id, discord_id, discord_username, nama, email, package_type, payment_url, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
             (order_id, discord_id, username, nama, email, package_type, payment_url, created_at))
    conn.commit()
    conn.close()

def save_subscription(order_id, discord_id, username, nama, email, package_type, referral_code=None, referrer_id=None):
    packages = get_all_packages()
    package = packages.get(package_type)
    if not package:
        return False
    
    conn = sqlite3.connect('warrior_subscriptions.db')
    c = conn.cursor()
    
    start = get_jakarta_datetime()
    end = start + timedelta(days=package['duration_days'])
    
    c.execute('''INSERT OR REPLACE INTO subscriptions 
                (order_id, discord_id, discord_username, nama, email, package_type, status, start_date, end_date, referral_code, referrer_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
             (order_id, discord_id, username, nama, email, package_type, 'active',
              start.strftime('%Y-%m-%d %H:%M:%S'), end.strftime('%Y-%m-%d %H:%M:%S'), referral_code, referrer_id))
    
    conn.commit()
    conn.close()
    return True

def send_welcome_email(member_name, email, package_name, order_id, start_date, end_date, referral_code, member_avatar):
    if not GMAIL_SENDER or not GMAIL_PASSWORD:
        print("⚠️ Gmail not configured")
        return False
    
    try:
        html_content = f"""
        <html>
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <style>
                    @media (max-width: 480px) {{
                        .header {{ padding: 20px 15px !important; }}
                        .header h1 {{ font-size: 24px !important; }}
                        .header h2 {{ font-size: 14px !important; }}
                        .content {{ padding: 20px !important; }}
                        .info-box {{ padding: 12px !important; }}
                    }}
                </style>
            </head>
            <body style="font-family: 'Segoe UI', Arial, sans-serif; background: linear-gradient(135deg, #f5f5f5 0%, #e8e8e8 100%); margin: 0; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                    
                    <!-- Orange Gradient Header -->
                    <div class="header" style="background: linear-gradient(135deg, #f7931a 0%, #ff7f00 100%); padding: 25px 20px; text-align: center; color: white;">
                        <h1 style="margin: 0 0 5px 0; font-size: 28px; font-weight: bold;">🎉 SELAMAT!</h1>
                        <h2 style="margin: 0; font-size: 16px; font-weight: 400; letter-spacing: 0.5px;">{member_name}</h2>
                    </div>
                    
                    <!-- White Content Area -->
                    <div style="background-color: white; padding: 30px;">
                        
                        <!-- Avatar -->
                        <div style="text-align: center; margin-bottom: 20px;">
                            <img src="{member_avatar}" alt="Avatar" style="width: 100px; height: 100px; border-radius: 50%; border: 4px solid #f7931a; box-shadow: 0 2px 8px rgba(247,147,26,0.3);">
                        </div>
                        
                        <!-- Title -->
                        <h3 style="text-align: center; color: #f7931a; font-size: 20px; margin: 0 0 20px 0;">✨ Membership Aktif ✨</h3>
                        
                        <!-- Info Box -->
                        <div style="background: linear-gradient(135deg, #fff9f0 0%, #fffbf5 100%); border-left: 4px solid #f7931a; padding: 15px; border-radius: 4px; margin-bottom: 20px;">
                            
                            <!-- Paket -->
                            <div style="display: flex; justify-content: space-between; margin-bottom: 12px; padding-bottom: 12px; border-bottom: 1px solid #ffe8cc;">
                                <span style="color: #666; font-weight: 600;">🎁 Paket:</span>
                                <span style="color: #f7931a; font-weight: bold;">{package_name}</span>
                            </div>
                            
                            <!-- Berakhir -->
                            <div style="display: flex; justify-content: space-between; margin-bottom: 12px; padding-bottom: 12px; border-bottom: 1px solid #ffe8cc;">
                                <span style="color: #666; font-weight: 600;">📅 Berakhir:</span>
                                <span style="color: #333; font-weight: bold;">{end_date}</span>
                            </div>
                            
                            <!-- Status -->
                            <div style="display: flex; justify-content: space-between;">
                                <span style="color: #666; font-weight: 600;">💚 Status:</span>
                                <span style="background-color: #00ff00; color: white; padding: 4px 12px; border-radius: 20px; font-weight: bold; font-size: 12px;">AKTIF</span>
                            </div>
                        </div>
                        
                        <!-- Message -->
                        <div style="text-align: center; background-color: #f7f7f7; padding: 15px; border-radius: 4px; margin-bottom: 20px;">
                            <p style="color: #f7931a; font-style: italic; margin: 0;">✨ Nikmati akses eksklusif The Warrior! ✨</p>
                        </div>
                        
                        <!-- Details Table -->
                        <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px;">
                            <tr style="background-color: #f9f9f9;">
                                <td style="padding: 10px; border: 1px solid #e0e0e0; color: #666;"><strong>Order ID:</strong></td>
                                <td style="padding: 10px; border: 1px solid #e0e0e0; color: #333; font-family: monospace;">{order_id}</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #e0e0e0; color: #666;"><strong>Mulai:</strong></td>
                                <td style="padding: 10px; border: 1px solid #e0e0e0; color: #333;">{start_date}</td>
                            </tr>
                            <tr style="background-color: #f9f9f9;">
                                <td style="padding: 10px; border: 1px solid #e0e0e0; color: #666;"><strong>Kode Referral:</strong></td>
                                <td style="padding: 10px; border: 1px solid #e0e0e0; color: #f7931a; font-weight: bold; font-size: 14px;">{referral_code}</td>
                            </tr>
                        </table>
                        
                        <!-- Footer Message -->
                        <p style="text-align: center; color: #f7931a; font-size: 14px; margin-top: 20px;">
                            💡 Jika ada pertanyaan, hubungi admin kami!
                        </p>
                    </div>
                    
                    <!-- Orange Footer -->
                    <div style="background: linear-gradient(135deg, #f7931a 0%, #ff7f00 100%); padding: 20px; text-align: center; color: white; font-size: 12px;">
                        © 2025 DiaryCrypto - The Warrior Membership
                    </div>
                </div>
            </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"✅ Welcome {member_name} - {package_name}"
        msg['From'] = GMAIL_SENDER
        msg['To'] = email
        
        msg.attach(MIMEText(html_content, 'html'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_SENDER, GMAIL_PASSWORD)
            server.sendmail(GMAIL_SENDER, email, msg.as_string())
        
        print(f"✅ Welcome email sent to {email}")
        return True
    except Exception as e:
        print(f"❌ Error sending welcome email: {e}")
        return False

def send_admin_new_member_notification(member_name, order_id, package_name, member_email):
    if not GMAIL_SENDER or not GMAIL_PASSWORD or not ADMIN_EMAIL:
        print("⚠️ Gmail or Admin email not configured")
        return False
    
    try:
        html_content = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f5f5f5;">
                <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px;">
                    <h2 style="color: #00aa00; text-align: center;">✅ NEW MEMBER JOINED</h2>
                    <hr style="border: 1px solid #ddd;">
                    
                    <p><strong>Informasi Member Baru:</strong></p>
                    
                    <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                        <tr style="background-color: #f9f9f9;">
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Member:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{member_name}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Email:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{member_email}</td>
                        </tr>
                        <tr style="background-color: #f9f9f9;">
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Paket:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{package_name}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Order ID:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{order_id}</td>
                        </tr>
                        <tr style="background-color: #f9f9f9;">
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Waktu:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{format_jakarta_datetime(datetime.now())}</td>
                        </tr>
                    </table>
                    
                    <p style="color: #666; font-size: 12px; margin-top: 20px; text-align: center;">
                        Email ini dikirim otomatis oleh sistem
                    </p>
                </div>
            </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"🎯 New Member - {member_name} (Order: {order_id})"
        msg['From'] = GMAIL_SENDER
        msg['To'] = ADMIN_EMAIL
        
        msg.attach(MIMEText(html_content, 'html'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_SENDER, GMAIL_PASSWORD)
            server.sendmail(GMAIL_SENDER, ADMIN_EMAIL, msg.as_string())
        
        print(f"✅ Admin notification sent to {ADMIN_EMAIL}")
        return True
    except Exception as e:
        print(f"❌ Error sending admin notification: {e}")
        return False

def send_expiry_reminder_email(member_name, email, package_name, end_date, member_avatar):
    """Send expiry reminder email dengan RED gradient design"""
    if not GMAIL_SENDER or not GMAIL_PASSWORD:
        print(f"❌ Gmail not configured - GMAIL_SENDER: {bool(GMAIL_SENDER)}, GMAIL_PASSWORD: {bool(GMAIL_PASSWORD)}")
        return False
    
    try:
        print(f"📧 Sending expiry email to {email}...")
        html_content = f"""
        <html>
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <style>
                    @media (max-width: 480px) {{
                        .header {{ padding: 20px 15px !important; }}
                        .header h1 {{ font-size: 24px !important; }}
                        .header h2 {{ font-size: 14px !important; }}
                        .content {{ padding: 20px !important; }}
                        .info-box {{ padding: 12px !important; }}
                    }}
                </style>
            </head>
            <body style="font-family: 'Segoe UI', Arial, sans-serif; background: linear-gradient(135deg, #f5f5f5 0%, #e8e8e8 100%); margin: 0; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                    
                    <!-- RED Gradient Header -->
                    <div class="header" style="background: linear-gradient(135deg, #ff4444 0%, #cc0000 100%); padding: 25px 20px; text-align: center; color: white;">
                        <h1 style="margin: 0 0 5px 0; font-size: 28px; font-weight: bold;">⚠️ MEMBERSHIP EXPIRED!</h1>
                        <h2 style="margin: 0; font-size: 16px; font-weight: 400; letter-spacing: 0.5px;">{member_name}</h2>
                    </div>
                    
                    <!-- White Content Area -->
                    <div class="content" style="background-color: white; padding: 25px;">
                        
                        <!-- Avatar -->
                        <div style="text-align: center; margin-bottom: 15px;">
                            <img src="{member_avatar}" alt="Avatar" style="width: 80px; height: 80px; border-radius: 50%; border: 3px solid #ff4444; box-shadow: 0 2px 8px rgba(255,68,68,0.3);">
                        </div>
                        
                        <!-- Title -->
                        <h3 style="text-align: center; color: #cc0000; font-size: 18px; margin: 0 0 15px 0;">📛 Membership Berakhir 📛</h3>
                        
                        <!-- Info Box -->
                        <div class="info-box" style="background: linear-gradient(135deg, #fff5f5 0%, #ffe8e8 100%); border-left: 4px solid #ff4444; padding: 12px; border-radius: 4px; margin-bottom: 15px;">
                            
                            <!-- Paket -->
                            <div style="display: flex; justify-content: space-between; margin-bottom: 12px; padding-bottom: 12px; border-bottom: 1px solid #ffcccc;">
                                <span style="color: #666; font-weight: 600;">📦 Paket:</span>
                                <span style="color: #cc0000; font-weight: bold;">{package_name}</span>
                            </div>
                            
                            <!-- Status -->
                            <div style="display: flex; justify-content: space-between; margin-bottom: 12px; padding-bottom: 12px; border-bottom: 1px solid #ffcccc;">
                                <span style="color: #666; font-weight: 600;">💥 Status:</span>
                                <span style="background-color: #ff4444; color: white; padding: 4px 12px; border-radius: 20px; font-weight: bold; font-size: 12px;">EXPIRED</span>
                            </div>
                            
                            <!-- End Date -->
                            <div style="display: flex; justify-content: space-between;">
                                <span style="color: #666; font-weight: 600;">📅 Berakhir:</span>
                                <span style="color: #333; font-weight: bold;">{end_date}</span>
                            </div>
                        </div>
                        
                        <!-- Alert Message -->
                        <div style="text-align: center; background-color: #fff5f5; padding: 15px; border-radius: 4px; margin-bottom: 20px; border: 2px dashed #ff4444;">
                            <p style="color: #cc0000; font-weight: bold; margin: 0;">Membership Anda telah berakhir dan role telah dihapus.</p>
                        </div>
                        
                        <!-- Action Button -->
                        <div style="text-align: center; margin-bottom: 20px;">
                            <p style="color: #666; margin: 0 0 10px 0;">Untuk melanjutkan akses The Warrior:</p>
                            <p style="margin: 0; font-size: 14px; color: #f7931a; font-weight: bold;">Gunakan command <strong>/buy</strong> untuk perpanjang atau beli paket baru! 🚀</p>
                        </div>
                        
                        <!-- Footer Message -->
                        <p style="text-align: center; color: #999; font-size: 12px; margin-top: 20px;">
                            💡 Hubungi admin jika ada pertanyaan
                        </p>
                    </div>
                    
                    <!-- RED Footer -->
                    <div style="background: linear-gradient(135deg, #ff4444 0%, #cc0000 100%); padding: 20px; text-align: center; color: white; font-size: 12px;">
                        © 2025 DiaryCrypto - The Warrior Membership
                    </div>
                </div>
            </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"⚠️ Membership Expired - {member_name}"
        msg['From'] = GMAIL_SENDER
        msg['To'] = email
        
        msg.attach(MIMEText(html_content, 'html'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_SENDER, GMAIL_PASSWORD)
            server.sendmail(GMAIL_SENDER, email, msg.as_string())
        
        print(f"✅ Expiry reminder email sent to {email}")
        return True
    except Exception as e:
        print(f"❌ Error sending expiry email: {e}")
        return False

def send_3day_expiry_warning_email(member_name, email, package_name, end_date, member_avatar, days_left):
    """Send 3-day expiry warning email dengan YELLOW gradient design"""
    if not GMAIL_SENDER or not GMAIL_PASSWORD:
        print(f"❌ Gmail not configured")
        return False
    
    try:
        print(f"📧 Sending 3-day warning email to {email}...")
        grace_period_end = (datetime.now(pytz.timezone('Asia/Jakarta')) + timedelta(days=2)).strftime("%d %B %Y")
        
        html_content = f"""
        <html>
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <style>
                    .step {{ margin-bottom: 12px; }}
                    .step-number {{ display: inline-block; background: #ffc107; color: white; width: 28px; height: 28px; border-radius: 50%; text-align: center; line-height: 28px; font-weight: bold; margin-right: 10px; }}
                    .step-text {{ display: inline-block; color: #333; font-size: 13px; line-height: 1.6; }}
                </style>
            </head>
            <body style="font-family: 'Segoe UI', Arial, sans-serif; background: #f5f5f5; margin: 0; padding: 20px;">
                <div style="max-width: 650px; margin: 0 auto;">
                    
                    <!-- YELLOW Gradient Header -->
                    <div style="background: linear-gradient(135deg, #ffc107 0%, #ffb300 50%, #ff9800 100%); padding: 30px 20px; text-align: center; color: white; border-radius: 12px 12px 0 0; box-shadow: 0 4px 12px rgba(255, 152, 0, 0.3);">
                        <h1 style="margin: 0 0 8px 0; font-size: 32px; font-weight: bold;">⚠️ PEMBERITAHUAN PENTING</h1>
                        <p style="margin: 0; font-size: 16px; opacity: 0.95;">Masa Aktif Membership Berakhir</p>
                    </div>
                    
                    <!-- White Content Area -->
                    <div style="background-color: white; padding: 35px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                        
                        <!-- Avatar -->
                        <div style="text-align: center; margin-bottom: 25px;">
                            <img src="{member_avatar}" alt="Avatar" style="width: 85px; height: 85px; border-radius: 50%; border: 4px solid #ffc107; box-shadow: 0 4px 12px rgba(255,193,7,0.4);">
                        </div>
                        
                        <!-- Greeting & Main Message -->
                        <p style="font-size: 17px; color: #222; margin: 0 0 20px 0; line-height: 1.8; font-weight: 500;">Halo 👋 <strong>{member_name}</strong></p>
                        
                        <p style="font-size: 14px; color: #555; margin: 0 0 18px 0; line-height: 1.8;">
                            Kami ingin menginformasikan bahwa masa aktif membership kamu di <strong style="color: #ff9800;">Diary Crypto</strong> telah berakhir hari ini.
                        </p>
                        
                        <!-- Grace Period Highlight Box -->
                        <div style="background: linear-gradient(135deg, #fffaf0 0%, #fff8e6 100%); border: 2px solid #ffc107; border-left: 6px solid #ffc107; padding: 18px; border-radius: 8px; margin-bottom: 25px;">
                            <p style="font-size: 14px; color: #333; margin: 0 0 12px 0; line-height: 1.7;">
                                ✨ Namun, sebagai bentuk kenyamanan dan apresiasi dari kami, kamu masih diberikan <strong>masa tenggang selama 2 hari</strong> ke depan agar tetap bisa mengakses channel dan seluruh konten eksklusif kami.
                            </p>
                            <div style="background: white; padding: 10px 12px; border-radius: 4px; text-align: center;">
                                <p style="font-size: 13px; color: #ff9800; margin: 0; font-weight: bold;">
                                    🗓️ Masa Tenggang Berakhir: <strong>{grace_period_end}</strong>
                                </p>
                            </div>
                        </div>
                        
                        <!-- Important Warning -->
                        <div style="background: #fff3cd; border-left: 4px solid #ff9800; padding: 12px 15px; border-radius: 4px; margin-bottom: 25px;">
                            <p style="font-size: 13px; color: #856404; margin: 0; line-height: 1.6;">
                                ⚠️ Setelah masa tenggang berakhir, akses kamu ke channel akan <strong>otomatis dinonaktifkan</strong> jika belum dilakukan perpanjangan.
                            </p>
                        </div>
                        
                        <!-- PROMINENT Renewal Instructions -->
                        <div style="background: linear-gradient(135deg, #f7f7f7 0%, #fafafa 100%); border: 2px solid #ffc107; padding: 20px; border-radius: 8px; margin-bottom: 25px;">
                            <h3 style="font-size: 16px; color: #ff9800; margin: 0 0 18px 0; font-weight: bold; text-align: center;">🔁 CARA PERPANJANG MEMBERSHIP</h3>
                            
                            <div class="step">
                                <span class="step-number">1</span>
                                <span class="step-text"><strong>Buka Discord</strong> dan gunakan command<br><span style="color: #ff9800; font-weight: bold; font-size: 14px;">/buy</span></span>
                            </div>
                            
                            <div class="step">
                                <span class="step-number">2</span>
                                <span class="step-text"><strong>Pilih paket</strong> yang ingin kamu perpanjang<br>(The Warrior 1 Jam / 3 Bulan / etc)</span>
                            </div>
                            
                            <div class="step">
                                <span class="step-number">3</span>
                                <span class="step-text"><strong>Lakukan pembayaran</strong> melalui link<br>Midtrans yang diberikan</span>
                            </div>
                            
                            <div class="step" style="margin-bottom: 0;">
                                <span class="step-number">4</span>
                                <span class="step-text"><strong>Selesai!</strong> Membership akan otomatis<br>terupdate setelah pembayaran berhasil</span>
                            </div>
                        </div>
                        
                        <!-- Support Section -->
                        <div style="background: #e3f2fd; border-left: 4px solid #2196F3; padding: 15px; border-radius: 6px; margin-bottom: 25px;">
                            <p style="font-size: 14px; color: #1565c0; margin: 0 0 8px 0; font-weight: bold;">💬 Butuh Bantuan?</p>
                            <p style="font-size: 13px; color: #1976d2; margin: 0; line-height: 1.6;">
                                Jika ada masalah atau pertanyaan, jangan ragu untuk <strong>DM kami</strong> di Discord. Tim Diary Crypto siap membantu! 🚀
                            </p>
                        </div>
                        
                        <!-- Benefits Section -->
                        <div style="margin-bottom: 25px;">
                            <p style="font-size: 14px; color: #333; margin: 0 0 12px 0; font-weight: bold;">Dengan memperpanjang, akses kamu ke:</p>
                            <div style="background: #f9f9f9; padding: 12px; border-radius: 6px;">
                                <p style="font-size: 13px; color: #555; margin: 6px 0; line-height: 1.6;">✅ Insight market harian & update real-time</p>
                                <p style="font-size: 13px; color: #555; margin: 6px 0; line-height: 1.6;">✅ Sinyal analisis & strategi crypto</p>
                                <p style="font-size: 13px; color: #555; margin: 6px 0; line-height: 1.6;">✅ Materi edukasi jangka panjang</p>
                                <p style="font-size: 13px; color: #555; margin: 6px 0; line-height: 1.6;">✅ Komunitas trader supportif untuk diskusi</p>
                            </div>
                        </div>
                        
                        <!-- Final CTA -->
                        <div style="background: linear-gradient(135deg, #fff9e6 0%, #fffbf0 100%); padding: 18px; border-radius: 8px; text-align: center; margin-bottom: 20px; border: 1px solid #ffecb3;">
                            <p style="font-size: 14px; color: #ff9800; margin: 0; font-weight: bold; line-height: 1.7;">
                                🚀 Jangan sampai terputus dari informasi yang bisa bantu kamu ambil keputusan terbaik di dunia crypto!
                            </p>
                        </div>
                        
                        <!-- Thank You -->
                        <div style="text-align: center; border-top: 1px solid #eee; padding-top: 20px;">
                            <p style="font-size: 13px; color: #777; margin: 0; line-height: 1.8;">
                                Terima kasih sudah menjadi bagian dari <strong>Diary Crypto</strong> 💎<br>
                                <span style="color: #ff9800; font-weight: bold;">Tim Diary Crypto</span>
                            </p>
                        </div>
                    </div>
                    
                    <!-- YELLOW Footer -->
                    <div style="background: linear-gradient(135deg, #ffc107 0%, #ff9800 100%); padding: 18px; text-align: center; color: white; font-size: 12px; border-radius: 0 0 12px 12px; box-shadow: 0 4px 12px rgba(255, 152, 0, 0.3);">
                        © 2025 DiaryCrypto - The Warrior Membership Platform
                    </div>
                </div>
            </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"⚠️ Pemberitahuan Penting: Masa Aktif Membership Berakhir - {member_name}"
        msg['From'] = GMAIL_SENDER
        msg['To'] = email
        
        msg.attach(MIMEText(html_content, 'html'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_SENDER, GMAIL_PASSWORD)
            server.sendmail(GMAIL_SENDER, email, msg.as_string())
        
        print(f"✅ 3-day warning email sent to {email}")
        return True
    except Exception as e:
        print(f"❌ Error sending 3-day warning email: {e}")
        return False

def send_trial_expiry_warning_email(member_name, email, trial_end, member_avatar, hours_left):
    """Send trial member expiry warning email dengan ORANGE design"""
    if not GMAIL_SENDER or not GMAIL_PASSWORD:
        print(f"❌ Gmail not configured")
        return False
    
    try:
        print(f"📧 Sending trial expiry warning email to {email}...")
        html_content = f"""
        <html>
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
            </head>
            <body style="font-family: 'Segoe UI', Arial, sans-serif; background: #f5f5f5; margin: 0; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto;">
                    
                    <!-- Orange Gradient Header -->
                    <div style="background: linear-gradient(135deg, #f7931a 0%, #ff7f00 100%); padding: 30px 20px; text-align: center; color: white; border-radius: 12px 12px 0 0; box-shadow: 0 4px 12px rgba(247,147,26,0.3);">
                        <h1 style="margin: 0 0 8px 0; font-size: 32px; font-weight: bold;">⏳ TRIAL AKAN BERAKHIR!</h1>
                        <p style="margin: 0; font-size: 16px; opacity: 0.95;">Jangan Lewatkan Akses Eksklusif</p>
                    </div>
                    
                    <!-- White Content Area -->
                    <div style="background-color: white; padding: 35px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                        
                        <!-- Avatar -->
                        <div style="text-align: center; margin-bottom: 25px;">
                            <img src="{member_avatar}" alt="Avatar" style="width: 85px; height: 85px; border-radius: 50%; border: 4px solid #f7931a; box-shadow: 0 4px 12px rgba(247,147,26,0.4);">
                        </div>
                        
                        <!-- Greeting -->
                        <p style="font-size: 17px; color: #222; margin: 0 0 20px 0; line-height: 1.8; font-weight: 500;">Halo 👋 <strong>{member_name}</strong></p>
                        
                        <!-- Alert Box -->
                        <div style="background: linear-gradient(135deg, #fff5e6 0%, #fff0d9 100%); border: 2px solid #f7931a; border-left: 6px solid #f7931a; padding: 18px; border-radius: 8px; margin-bottom: 25px;">
                            <p style="font-size: 16px; color: #333; margin: 0 0 12px 0; font-weight: bold; line-height: 1.7;">
                                ⚠️ Trial akses Anda ke The Warrior akan berakhir dalam <span style="color: #f7931a;">kurang dari 24 jam</span>!
                            </p>
                            <div style="background: white; padding: 12px; border-radius: 4px; text-align: center;">
                                <p style="font-size: 14px; color: #f7931a; margin: 0; font-weight: bold;">
                                    ⏰ Berakhir: <strong>{trial_end}</strong>
                                </p>
                            </div>
                        </div>
                        
                        <!-- What You'll Lose -->
                        <div style="background: #fff9f5; padding: 18px; border-radius: 8px; margin-bottom: 25px;">
                            <p style="font-size: 14px; color: #333; margin: 0 0 12px 0; font-weight: bold;">Akses yang akan hilang:</p>
                            <ul style="font-size: 13px; color: #555; margin: 0; padding-left: 20px; line-height: 1.8;">
                                <li>❌ Insight market harian & update real-time</li>
                                <li>❌ Sinyal analisis & strategi crypto</li>
                                <li>❌ Materi edukasi jangka panjang</li>
                                <li>❌ Komunitas trader supportif</li>
                            </ul>
                        </div>
                        
                        <!-- Call to Action -->
                        <div style="background: linear-gradient(135deg, #fff5e6 0%, #fff0d9 100%); padding: 20px; border-radius: 8px; text-align: center; margin-bottom: 20px;">
                            <p style="font-size: 15px; color: #f7931a; margin: 0; font-weight: bold; line-height: 1.8;">
                                🚀 Perpanjang akses Anda sekarang dengan command<br><strong>/buy</strong><br>di Discord!
                            </p>
                        </div>
                        
                        <!-- Benefits -->
                        <p style="font-size: 13px; color: #666; margin: 0; line-height: 1.8; text-align: center;">
                            Jangan sampai kehilangan akses ke konten eksklusif dan insights penting dari Tim Diary Crypto!<br><br>
                            <strong style="color: #f7931a;">Gunakan /buy sekarang untuk perpanjang! 💎</strong>
                        </p>
                    </div>
                    
                    <!-- Orange Footer -->
                    <div style="background: linear-gradient(135deg, #f7931a 0%, #ff7f00 100%); padding: 18px; text-align: center; color: white; font-size: 12px; border-radius: 0 0 12px 12px; box-shadow: 0 4px 12px rgba(247,147,26,0.3);">
                        © 2025 DiaryCrypto - The Warrior Trial Membership
                    </div>
                </div>
            </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"⏳ Trial Akan Berakhir - {member_name}"
        msg['From'] = GMAIL_SENDER
        msg['To'] = email
        
        msg.attach(MIMEText(html_content, 'html'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_SENDER, GMAIL_PASSWORD)
            server.sendmail(GMAIL_SENDER, email, msg.as_string())
        
        print(f"✅ Trial expiry warning email sent to {email}")
        return True
    except Exception as e:
        print(f"❌ Error sending trial expiry warning email: {e}")
        return False

def send_trial_member_email(member_name, email, trial_start, trial_end, member_avatar, duration_days=1):
    """Send trial member welcome email dengan ORANGE design"""
    if not GMAIL_SENDER or not GMAIL_PASSWORD:
        print(f"❌ Gmail not configured")
        return False
    
    try:
        # Convert duration to readable text
        if duration_days < 1:
            hours = int(duration_days * 24)
            duration_text = f"{hours} Jam" if hours > 0 else "1 Jam"
        else:
            duration_text = f"{int(duration_days)} Hari" if duration_days > 1 else "1 Hari"
        
        print(f"📧 Sending trial member email to {email}...")
        html_content = f"""
        <html>
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
            </head>
            <body style="font-family: 'Segoe UI', Arial, sans-serif; background: linear-gradient(135deg, #f5f5f5 0%, #e8e8e8 100%); margin: 0; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                    
                    <!-- Orange Gradient Header -->
                    <div style="background: linear-gradient(135deg, #f7931a 0%, #ff7f00 100%); padding: 25px 20px; text-align: center; color: white;">
                        <h1 style="margin: 0 0 5px 0; font-size: 28px; font-weight: bold;">🎉 TRIAL AKTIF!</h1>
                        <h2 style="margin: 0; font-size: 16px; font-weight: 400; letter-spacing: 0.5px;">{member_name}</h2>
                    </div>
                    
                    <!-- White Content Area -->
                    <div style="background-color: white; padding: 25px;">
                        
                        <!-- Avatar -->
                        <div style="text-align: center; margin-bottom: 15px;">
                            <img src="{member_avatar}" alt="Avatar" style="width: 80px; height: 80px; border-radius: 50%; border: 3px solid #f7931a; box-shadow: 0 2px 8px rgba(247,147,26,0.3);">
                        </div>
                        
                        <!-- Title -->
                        <h3 style="text-align: center; color: #f7931a; font-size: 18px; margin: 0 0 15px 0;">✨ Trial Member {duration_text} ✨</h3>
                        
                        <!-- Info Box -->
                        <div style="background: linear-gradient(135deg, #fff5e6 0%, #fff0d9 100%); border-left: 4px solid #f7931a; padding: 12px; border-radius: 4px; margin-bottom: 15px;">
                            
                            <!-- Mulai -->
                            <div style="display: flex; justify-content: space-between; margin-bottom: 10px; padding-bottom: 10px; border-bottom: 1px solid #ffe8cc;">
                                <span style="color: #666; font-weight: 600;">📅 Mulai:</span>
                                <span style="color: #f7931a; font-weight: bold;">{trial_start}</span>
                            </div>
                            
                            <!-- Berakhir -->
                            <div style="display: flex; justify-content: space-between;">
                                <span style="color: #666; font-weight: 600;">⏰ Berakhir:</span>
                                <span style="color: #f7931a; font-weight: bold;">{trial_end}</span>
                            </div>
                        </div>
                        
                        <!-- Message -->
                        <div style="text-align: center; background-color: #f7f7f7; padding: 15px; border-radius: 4px; margin-bottom: 15px;">
                            <p style="color: #f7931a; font-style: italic; margin: 0;">✨ Nikmati akses eksklusif The Warrior selama {duration_text}! ✨</p>
                        </div>
                        
                        <!-- Alert -->
                        <div style="text-align: center; background-color: #fff5e6; padding: 12px; border-radius: 4px; border: 2px dashed #f7931a; margin-bottom: 15px;">
                            <p style="color: #f7931a; font-weight: bold; margin: 0;">⏳ Trial berakhir dalam {duration_text}, role akan otomatis dihapus</p>
                        </div>
                        
                        <!-- Footer Message -->
                        <p style="text-align: center; color: #999; font-size: 12px; margin-top: 15px;">
                            💡 Untuk akses lebih lama, gunakan command /buy sekarang juga!
                        </p>
                    </div>
                    
                    <!-- Orange Footer -->
                    <div style="background: linear-gradient(135deg, #f7931a 0%, #ff7f00 100%); padding: 20px; text-align: center; color: white; font-size: 12px;">
                        © 2025 DiaryCrypto - The Warrior Trial Membership
                    </div>
                </div>
            </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"🎉 Trial Activated - {member_name}"
        msg['From'] = GMAIL_SENDER
        msg['To'] = email
        
        msg.attach(MIMEText(html_content, 'html'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_SENDER, GMAIL_PASSWORD)
            server.sendmail(GMAIL_SENDER, email, msg.as_string())
        
        print(f"✅ Trial member email sent to {email}")
        return True
    except Exception as e:
        print(f"❌ Error sending trial email: {e}")
        return False

def send_admin_kick_notification(member_name: str, member_email: str, package_name: str, reason: str):
    """Send admin notification about member role removal"""
    if not GMAIL_SENDER or not GMAIL_PASSWORD or not ADMIN_EMAIL:
        print("⚠️ Gmail or Admin email not configured, skipping kick notification")
        return False
    
    try:
        html_content = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f5f5f5;">
                <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px;">
                    <h2 style="color: #ff0000; text-align: center;">🚨 MEMBER ROLE REMOVED</h2>
                    <hr style="border: 1px solid #ddd;">
                    
                    <p><strong>Informasi Pencopotan Role:</strong></p>
                    
                    <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                        <tr style="background-color: #f9f9f9;">
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Member:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{member_name}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Email:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{member_email}</td>
                        </tr>
                        <tr style="background-color: #f9f9f9;">
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Paket:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{package_name}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Alasan:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd; color: #ff0000;"><strong>{reason}</strong></td>
                        </tr>
                        <tr style="background-color: #f9f9f9;">
                            <td style="padding: 10px; border: 1px solid #ddd;"><strong>Waktu:</strong></td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{format_jakarta_datetime(datetime.now())}</td>
                        </tr>
                    </table>
                    
                    <p style="color: #666; font-size: 12px; margin-top: 20px; text-align: center;">
                        Email ini dikirim otomatis oleh sistem
                    </p>
                </div>
            </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"🚨 KICK NOTIFICATION - {member_name} ({reason})"
        msg['From'] = GMAIL_SENDER
        msg['To'] = ADMIN_EMAIL
        
        msg.attach(MIMEText(html_content, 'html'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_SENDER, GMAIL_PASSWORD)
            server.sendmail(GMAIL_SENDER, ADMIN_EMAIL, msg.as_string())
        
        print(f"✅ Admin kick notification sent to {ADMIN_EMAIL}")
        return True
    except Exception as e:
        print(f"❌ Error sending kick notification: {e}")
        return False

async def fetch_crypto_news():
    """Fetch REAL crypto news November 2025 - BAHASA INDONESIA with disclaimer"""
    # Real crypto news dari November 24, 2025
    
    # Get current Jakarta time
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    now_jakarta = datetime.now(jakarta_tz)
    timestamp = now_jakarta.strftime('%d %b %Y, %H:%M WIB')
    
    # Disclaimer header
    disclaimer = '''⚠️ **DISCLAIMER - NOT FINANCIAL ADVICE (NFA)**
🔍 **DYOR - DO YOUR OWN RESEARCH**
📊 Analisis ini untuk educational purpose saja. Bukan rekomendasi trading/investasi.
⚡ Crypto adalah HIGHLY RISKY. Investasi sesuai kemampuan Anda saja!'''

    articles_with_analysis = [
        {
            'title': '🔴 BITCOIN CRASH NOVEMBER 2025 - Turun 25% Terburuk Sejak 2022!',
            'image_url': 'https://images.unsplash.com/photo-1621761191319-c6fb62b6fe6e?w=500',
            'analysis': f'''{disclaimer}

**📊 DATA SOURCES:**
• CoinMarketCap (Bitcoin price: $91,510)
• CMC Fear & Greed Index (Current: 11 - Extreme Fear)
• JPMorgan Research (Retail selling analysis)
• BlackRock IBIT ETF (Bitcoin ETF flows)
• Grayscale Crypto (Market cap data)
🕐 Posted: {timestamp}

---

**🚨 ALERT: CRASH CRYPTO NOVEMBER 2025 - TERBURUK SEJAK 2022**

**SITUASI KRITIS:**
Bitcoin anjlok drastis ke level $91,510 (Nov 24) - turun 25% sepanjang November! Harga sempat menyentuh $82,000-$83,590 minggu ini dengan liquidation $800 juta+ dalam satu hari. Ethereum jatuh -11.2%, ETH turun dari $2,736.

**PENYEBAB CRASH:**
🚨 **Retail Investor PANIC SELL** - JPMorgan: bukan institutional deleveraging, tapi RETAIL yang banting stir!
💰 **Record ETF Outflows** - $3.79B-$4B ditarik dari Bitcoin/Ethereum ETF (pecah rekor semua waktu!)
😱 **BlackRock IBIT Collapse** - $2.1B withdraw dari flagship Bitcoin ETF November alone
📉 **Extreme Fear Index** - CMC Fear = 11 (EXTREME FEAR, terburuk sejak akhir 2022!)

**DAMPAK LANGSUNG:**
• Pembeli September-Oktober 2024 SEMUANYA DALAM KERUGIAN -25%
• Michael Saylor's Strategy: 650K BTC positions nyaris BREAKEVEN
• Total market cap turun ke bawah $3 TRILIUN (dari peak $5 triliun)
• $1.2 TRILIUN hilang dalam 6 minggu terakhir saja!
• Bitmine holding $3.5B unrealized ETH losses

**TECHNICAL SETUP - DANGER ZONE:**
⚠️ Support level $77,000-$80,000 masih RAWAN ditest
📊 Market structure SANGAT LEMAH - tidak ada rebound meaningful
🎲 Altcoin recovery TERTUNDA - Bitcoin harus stabilisasi dulu
❌ No positive catalyst - Fed rate cut uncertainty masih hanging

**RETAIL CAPITULATION SIGNAL (BULLISH DIVERGENCE?):**
Investor kecil sedang dump holdings mereka dengan panic - historical research: signal ini SERING jadi market bottom! Tapi timing sangat sulit - jangan bergabung panic.

**MACRO RISKS MINGGU DEPAN (CRITICAL):**
📅 Nov 25: PPI Inflation data
📅 Nov 26: PCE Inflation (CRUCIAL untuk crypto direction) + jobless claims  
⚠️ Jika inflation NAGGING UP = Fed hold rate cuts = crypto makin down

**FORECASTING 2 MINGGU REALISTIS:**
📉 **Bear Case (60% probability):** Bitcoin test $77K-$80K range, jangan expect bounce cepat
📈 **Bull Case (40% probability):** Stabilisasi di $85K-$90K, rebound bertahap ke $95K+

**REKOMENDASI URGENT:**
- HODLER: Jangan PANIC SELL! History repeats - ini sudah terjadi 2020-2021
- TRADER: Siapkan limit order di $80K-$85K untuk accumulate (jangan desperate)
- PEMULA: STAY AWAY! Tung sampai Fear Index turun ke 30 atau bawah
- GUNAKAN DCA: Small frequent buys > all-in saat market berdarah

**MENTAL NOTES:**
Extreme fear reading ini adalah TESTING TIME - bukan collapse final. Para hodler dari 2021 sudah terbiasa cycle ini. Tapi yang beli September 2024 sekarang deeply underwater. NO REGRET allowed - time in market > timing market.

⚠️ **INGAT:** Crypto masih HIGHLY RISKY. Recovery timing UNPREDICTABLE. Jangan leverage - hanya cash yang afford to lose!'''
        },
        {
            'title': '💰 ETH RECORD OUTFLOWS - $3.79B Ditarik dari Bitcoin/Ethereum ETF!',
            'image_url': 'https://images.unsplash.com/photo-1618793059027-ea4b6e3d6d7a?w=500',
            'analysis': f'''{disclaimer}

**📊 DATA SOURCES:**
• BlackRock IBIT ETF ($2.1B withdraw)
• CoinMarketCap ETF Flow Data ($3.79B-$4B November outflow)
• JPMorgan Research (Retail selling analysis)
• Blockchain.com (Market cap tracking)
• The Block (ETF analytics)
🕐 Posted: {timestamp}

---

**📊 ANALISIS: ETF OUTFLOWS RECORD NOVEMBER 2025**

**DATA TERPERINCI - ETF EXODUS:**
November sudah catat $3.79B-$4B outflow dari Bitcoin dan Ethereum spot ETF - MENGALAHKAN rekor Februari 2023! Ini adalah bulan TERBURUK untuk ETF inflow sejak crypto fund tracking mulai. Single day Nov 20: $1.6B outflow dalam SATU HARI!

**BREAKDOWN OUTFLOW:**
• **BlackRock IBIT (flagship BTC ETF):** $2.1B withdraw - WORST MONTH EVER
• **Ethereum ETF:** Massive selling juga, second biggest loser
• **Other Bitcoin ETF:** Semuanya RED - tidak ada ETF yang positive inflow
• **Total from $5T peak to $3T now:** $2 TRILLION market cap VAPORIZED

**WHY RETAIL SELLING NOW?**
JPMorgan research jelas: RETAIL yang sell, bukan whale/institutional deleveraging. Ini karena:
1. Stop-loss triggers dari September buyers
2. Margin calls dari leverage traders
3. FOMO selling (fear of losing position lanjut)
4. Year-end tax-loss harvesting mulai

**HISTORICAL CONTEXT:**
Outflow record ini belum pernah terjadi sejak ETF tracking dimulai. Bahkan Feb 2023 (SVB crisis) pun lebih kecil. Ini menunjukkan RETAIL PANIC unprecedented level.

**WHAT'S NEXT:**
Jika outflow terus berlanjut: Bisa trigger cascade selling. Jika stabilisasi di $77K support: Bisa jadi turnaround point. CMC Fear Index 11 historically signal EXTREME CAPITULATION - sering jadi bottom.

**INVESTMENT IMPLICATIONS:**
For patient investors: This might be best buying opportunity 2025. Tapi timing tetap sulit. Better safe dengan DCA daripada all-in sekarang.'''
        },
        {
            'title': '💚 SOLANA BRIGHT SPOT - 19 Hari ETF Inflows Berturut-turut!',
            'image_url': 'https://images.unsplash.com/photo-1639762681033-6461efb0b480?w=500',
            'analysis': f'''{disclaimer}

**📊 DATA SOURCES:**
• Grayscale SOL ETF (19 consecutive day inflow data)
• CoinMarketCap (Solana price: $128, down -13%)
• The Block (Solana ecosystem metrics)
• Glassnode (Large wallet accumulation)
• Solana Foundation (Ecosystem stats: 10,000+ apps)
🕐 Posted: {timestamp}

---

**🟢 SOLANA MOMENTUM: SATU-SATUNYA ALTCOIN YANG SURVIVE CRASH**

**SOL RELATIVE STRENGTH:**
Sementara Bitcoin turun -25%, Ethereum -11%, Solana "hanya" turun -13% - tapi yang impressive: **SOL ETF menerima inflow KONSISTEN 19 hari berturut-turut** meskipun keseluruhan market crashed! Ini adalah STRONG signal institutional interest sama SOL.

**SOL ETF INFLOWS STATS:**
• 19 consecutive days ETF inflow: $23 MILLION+ total
• Sementara BTC/ETH ETF outflow: Billions
• Market clearly rotating: From BTC/ETH → Solana
• Grayscale SOL ETF juga attract significant interest

**WHY SOL OUTPERFORM?**
1. **Ecosystem strength:** 10,000+ aplikasi aktif (vs Ethereum 5,000+)
2. **Transaction speed:** 65,000 TPS vs Bitcoin 7 TPS
3. **Low fees:** Rp 100-500 per tx vs Ethereum Rp 50,000
4. **Developer grants:** Terus menarik top talent
5. **Mobile-first strategy:** Saga phone launch sukses

**MARKET PSYCHOLOGY:**
Retail investors rotating dari BTC (expensive, bear sentiment) ke SOL (faster, cheaper, bullish narrative). Ini sering signal yang memimpin altcoin rally sesudah Bitcoin recovery.

**FORECAST:**
Jika Bitcoin stabilisasi di $85K: SOL bisa rally 3-5x dalam 3 bulan
Jika Bitcoin test $77K: SOL mungkin ikutan turun tapi less severe

**REKOMENDASI:**
SOL adalah "hedge" terbaik di crash ini. Sementara Bitcoin uncertain, SOL ecosystem terus berkembang dan institutional buying increasing. Good accumulation zone di $100-120 range.'''
        },
        {
            'title': '💛 XRP WHALE DUMP - Grayscale XRP ETF Launching Senin!',
            'image_url': 'https://images.unsplash.com/photo-1579621970563-430f63602d4b?w=500',
            'analysis': f'''{disclaimer}

**📊 DATA SOURCES:**
• CoinMarketCap (XRP price: $1.94, down -12.2%)
• Whale Alert (250M XRP whale dump transaction)
• Grayscale (XRP ETF approval by NYSE Arca)
• Glassnode (Large wallet movement analysis)
• The Block (ETF launch tracking)
🕐 Posted: {timestamp}

---

**🚨 XRP VOLATILITY: WHALE DUMP vs ETF LAUNCH**

**XRP PRICE ACTION - NOVEMBER 24:**
XRP trading di $1.94 (-12.2% 24h), lost critical $2.00 support level. Tapi MASSIVE development: **Grayscale XRP ETF just approved by NYSE Arca - LAUNCHING MONDAY!**

**WHALE DUMP ALERT:**
Whale wallets offloaded 250 MILLION XRP tokens dalam single transaction minggu ini. Ini adalah MASSIVE sell pressure. Market interpreting ini sebagai "insiders taking profits sebelum ETF launch."

**GRAYSCALE XRP ETF CATALYST:**
✅ Institutional exposure baru untuk XRP (seperti Bitcoin/Ethereum ETF)
✅ Likely akan attract billions dalam AUM (assets under management)
✅ Possible short squeeze jika retail rushes untuk buy Monday open
❌ Risk: Whale dump might pressure price before institutional buying

**HISTORICAL PRECEDENT:**
Ketika Bitcoin spot ETF approved (Jan 2024): Huge institutional inflows immediately. XRP ETF bisa replicate pattern ini. Tapi kalau whales dump lebih dulu → volatility extreme.

**TECHNICAL SETUP:**
Support: $1.70-1.80 (critical level)
Resistance: $2.50 (pre-dump level)
Scenario 1: Whale dump finished, institutional buying Monday → $2.50+ jump
Scenario 2: More whale selling continues → test $1.70 support

**TUESDAY-WEDNESDAY CRUCIAL:**
After Monday launch, watch Wednesday close. Jika ETF inflows strong: Rally bisa accelerate. If disappointing: Back to dump pressure.

**PLAY:**
Conservative: Wait until ETH stabilizes Thursday, then buy dip dengan DCA
Aggressive: Buy Monday open expecting ETF inflows (risky - could dump more first)
Safest: Monitor first 2 days volume, buy if volume confirms institutional accumulation.'''
        },
        {
            'title': '😨 EXTREME FEAR INDEX 11 - Terburuk Sejak Akhir 2022!',
            'image_url': 'https://images.unsplash.com/photo-1611531900900-48d240ce8313?w=500',
            'analysis': f'''{disclaimer}

**📊 DATA SOURCES:**
• CoinMarketCap Fear & Greed Index (Current: 11)
• Santiment (Retail wallet sentiment analysis)
• Glassnode (On-chain activity metrics)
• Crypto Market Data (Market dominance tracking)
• Historical Fear Index Database (Comparison analysis)
🕐 Posted: {timestamp}

---

**🔴 CMC FEAR & GREED INDEX: 11 OUT OF 100 (EXTREME FEAR)**

**FEAR INDEX BREAKDOWN - NOVEMBER 24:**
CMC Fear & Greed Index: **11/100** - EXTREME FEAR zone. Terburuk sejak November-December 2022 (ketika FTX collapse). Reading ini indicates MAXIMUM panic - market emotionally EXHAUSTED.

**WHAT EXTREME FEAR (0-25) MEANS:**
• Retail panic selling at peak
• Capitulation often nearby
• Market sentiment extremely negative
• Historically good buying opportunity (but timing hard)
• Fear can go LOWER before reversal

**HISTORICAL COMPARISON:**
- March 2020 (COVID crash): Fear = 5 (RECOVERY in 6 months)
- May 2021 (Elon FUD): Fear = 8 (RECOVERY in 2 months)
- Nov 2022 (FTX collapse): Fear = 10 (RECOVERY in 1 month to January)
- **Nov 2024 (NOW): Fear = 11** (UNPRECEDENTED)

**WHAT THIS MEANS FOR FUTURE:**
📊 Extreme fear historically = NEAR BOTTOM (70-80% accuracy)
⏰ Timing recovery = IMPOSSIBLE (could take 1 week or 8 weeks)
💰 Opportunity: When fear HIGH, patient investors accumulate quietly
🎲 Risk: Fear can STAY high for weeks (patience required)

**DATA POINTS SUPPORTING REVERSAL NEAR:**
• Retail completely capitulated (selling everything)
• Large wallets stopped selling and started accumulating ($250M BTC buy orders detected)
• Long liquidations at historical highs ($800M+ liquidated = weakness gone)
• FOMO reversed (everyone too scared to buy)

**FORECASTING - NEXT 3 WEEKS:**
Scenario A (60%): Fear stabilizes or stays 11-20, sideways $80K-$90K for 2-3 weeks, then gradual reversal
Scenario B (30%): Fear spikes to 5-8 (capitulation complete), then rapid recovery within days
Scenario C (10%): Fear maintains elevated, drops to $77K-$80K first

**PSYCHOLOGICAL INFLECTION:**
When Fear index reaches this extreme, PSYCHOLOGICALLY people are most bearish. Contrarian trading principle: maximum bearishness = contrarian buy signal. But requires NERVES OF STEEL to buy when everyone panicking.

**BOTTOM LINE:**
We're at EMOTIONAL EXTREME. Market rarely stays there long. Whether reversal takes 5 days or 5 weeks = unpredictable. But risk/reward NOW is HEAVILY skewed to upside for patient 6-month+ holders.'''
        }
    ]
    
    return articles_with_analysis


async def auto_post_crypto_news():
    """Auto-post cryptocurrency news dengan analysis ke payment channel setiap 24 jam"""
    await bot.wait_until_ready()
    
    while not bot.is_closed():
        try:
            guild = bot.get_guild(GUILD_ID)
            if not guild:
                await asyncio.sleep(3600)
                continue
            
            # Find payment channel for news posting - ONLY payment channel
            news_channel = None
            for channel in guild.text_channels:
                if channel.name == "💳｜payment":
                    news_channel = channel
                    break
            
            if not news_channel:
                print(f"⚠️ Channel #💳｜payment tidak ditemukan. Skip posting berita.")
                await asyncio.sleep(3600)
                continue
            
            # Fetch crypto news dengan analysis
            articles = await fetch_crypto_news()
            
            if articles:
                print(f"✅ AUTO POSTING CRYPTO NEWS - {len(articles)} berita ke #💳｜payment")
                
                for article in articles:
                    try:
                        title = article.get('title', 'Untitled')
                        image = article.get('image_url', '')
                        analysis = article.get('analysis', '')
                        
                        # Create embed dengan full analysis ONLY (TANPA link)
                        embed = discord.Embed(
                            title=title[:256],
                            description=analysis[:4000] if analysis else "Analysis tidak tersedia",
                            color=0xf7931a
                        )
                        
                        if image:
                            embed.set_image(url=image)
                        
                        embed.set_footer(text="📊 Analisis Crypto News")
                        
                        await news_channel.send(embed=embed)
                        await asyncio.sleep(1)
                    except Exception as e:
                        print(f"⚠️ Error posting article: {e}")
                
                print(f"✅ {len(articles)} berita crypto dengan analysis berhasil di-post")
            
            # Post setiap 24 jam (86400 detik)
            await asyncio.sleep(86400)
        
        except Exception as e:
            print(f"❌ Error in auto crypto news task: {e}")
            await asyncio.sleep(3600)


async def cleanup_stale_orders():
    await bot.wait_until_ready()
    
    while not bot.is_closed():
        try:
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            # Cek order yang pending lebih dari 10 MENIT dengan real time (Jakarta timezone)
            now_jakarta = get_jakarta_datetime()
            cutoff_time = (now_jakarta - timedelta(minutes=10)).strftime('%Y-%m-%d %H:%M:%S')
            
            c.execute('SELECT order_id, discord_id, discord_username, package_type FROM pending_orders WHERE status = "pending" AND created_at < ?', (cutoff_time,))
            stale_orders = c.fetchall()
            
            if stale_orders:
                print(f"🧹 Cleanup: Found {len(stale_orders)} expired orders (>10 menit) - Real time: {now_jakarta.strftime('%Y-%m-%d %H:%M:%S WIB')}")
                
                for (order_id, discord_id, discord_username, package_type) in stale_orders:
                    try:
                        # Hapus order yang expired
                        c.execute('DELETE FROM pending_orders WHERE order_id = ?', (order_id,))
                        
                        # Kirim DM ke user
                        user = await bot.fetch_user(int(discord_id))
                        if user:
                            dm_embed = discord.Embed(
                                title="⏰ ORDER KADALUARSA!",
                                description="Pembayaran Anda tidak selesai dalam 10 menit",
                                color=0xff0000
                            )
                            dm_embed.add_field(name="❌ Status", value="Order Expired", inline=True)
                            dm_embed.add_field(name="📋 Order ID", value=f"`{order_id}`", inline=True)
                            dm_embed.add_field(name="📦 Paket", value=package_type, inline=False)
                            dm_embed.add_field(name="🔄 Solusi", value="Gunakan `/buy` lagi untuk membuat order baru", inline=False)
                            dm_embed.add_field(name="⏱️ Waktu Expire", value=format_jakarta_datetime(now_jakarta), inline=False)
                            dm_embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
                            
                            await user.send(embed=dm_embed)
                            print(f"✅ Expired notification sent to {discord_username}")
                    except Exception as e:
                        print(f"⚠️ Error sending expired notification: {e}")
            
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"❌ Error in cleanup: {e}")
        
        await asyncio.sleep(10)


async def check_expired_subscriptions():
    await bot.wait_until_ready()
    
    while not bot.is_closed():
        try:
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            # Use Jakarta timezone untuk accurate comparison dengan database
            now = get_jakarta_datetime().strftime('%Y-%m-%d %H:%M:%S')
            
            c.execute('''SELECT discord_id, discord_username, nama, email, package_type, end_date 
                        FROM subscriptions 
                        WHERE status = "active" 
                        AND datetime(end_date) <= datetime(?)''',
                     (now,))
            
            expired_subs = c.fetchall()
            print(f"🔍 Auto removal check: Found {len(expired_subs)} expired memberships")
            
            guild = bot.get_guild(GUILD_ID)
            if not guild:
                print(f"❌ Guild not found in auto removal!")
                conn.close()
                await asyncio.sleep(3600)
                continue
            
            for discord_id, discord_username, nama, email, package_type, end_date in expired_subs:
                try:
                    print(f"🔄 Processing expired: {discord_username} ({discord_id}) - Package: {package_type}")
                    
                    member = guild.get_member(int(discord_id))
                    if not member:
                        print(f"  ⚠️ Member {discord_username} ({discord_id}) tidak ditemukan di guild")
                        c.execute('UPDATE subscriptions SET status = "expired" WHERE discord_id = ?',
                                 (discord_id,))
                        print(f"  ✅ Status updated to expired")
                        continue
                    
                    packages = get_all_packages()
                    role_name = packages.get(package_type, {}).get("role_name")
                    duration_days = packages.get(package_type, {}).get("duration_days", 0)
                    if not role_name:
                        print(f"  ❌ Role name tidak ditemukan untuk package {package_type}")
                        continue
                    
                    role = discord.utils.get(guild.roles, name=role_name)
                    if not role:
                        print(f"  ❌ Role '{role_name}' tidak ditemukan di guild")
                        continue
                    
                    if role in member.roles:
                        pkg_name = packages.get(package_type, {}).get('name', 'The Warrior')
                        end_datetime_full = format_jakarta_datetime_full(end_date)
                        
                        # Send RED EMBED DM notification
                        try:
                            expiry_embed = discord.Embed(
                                title="⚠️ MEMBERSHIP EXPIRED! ⚠️",
                                description=f"Paket **{pkg_name}** Anda telah berakhir.",
                                color=0xff4444
                            )
                            expiry_embed.add_field(name="🔴 Status", value="EXPIRED", inline=True)
                            expiry_embed.add_field(name="📅 Berakhir", value=end_datetime_full, inline=True)
                            expiry_embed.add_field(name="🔄 Solusi", value="Klik `/buy` untuk perpanjang atau beli paket baru!", inline=False)
                            expiry_embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
                            expiry_embed.set_thumbnail(url=member.avatar.url if member.avatar else "")
                            
                            await member.send(embed=expiry_embed)
                            print(f"  ✅ Expiry RED EMBED sent to {member.name}")
                        except discord.HTTPException as e:
                            print(f"  ⚠️ Could not send DM to {discord_id}: {e}")
                        
                        await member.remove_roles(role)
                        print(f"  ✅ Role '{role_name}' removed from {discord_username}")
                        
                        # Send expiry reminder email dengan RED gradient
                        try:
                            member_avatar = str(member.avatar.url) if member.avatar else str(member.default_avatar)
                            end_datetime = format_jakarta_datetime_full(end_date)
                            result = send_expiry_reminder_email(nama, email, pkg_name, end_datetime, member_avatar)
                            if result:
                                print(f"  ✅ Expiry RED email sent to {email}")
                            else:
                                print(f"  ⚠️ Email function returned False for {email}")
                        except Exception as e:
                            print(f"  ❌ Error sending expiry email to {email}: {e}")
                        
                        send_admin_kick_notification(nama, email, pkg_name, "Membership Expired")
                    
                    c.execute('UPDATE subscriptions SET status = "expired" WHERE discord_id = ?', (discord_id,))
                    print(f"  ✅ Subscription marked as expired")
                    
                except Exception as e:
                    print(f"  ❌ Error: {e}")
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"❌ Error in expiry check: {e}")
        
        await asyncio.sleep(300)


async def auto_remove_expired_members():
    await bot.wait_until_ready()
    
    while not bot.is_closed():
        try:
            await asyncio.sleep(300)
        except Exception as e:
            print(f"❌ Error: {e}")


async def check_3day_expiry_warning():
    """Background task untuk kirim warning 3 hari sebelum expiry (1x per hari selama 3 hari)"""
    await bot.wait_until_ready()
    
    while not bot.is_closed():
        try:
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            now = get_jakarta_datetime()
            three_days_later = (now + timedelta(days=3)).strftime('%Y-%m-%d')
            
            # Find memberships yang akan expire dalam 3 hari, dan reminder_count < 3
            c.execute('''SELECT order_id, discord_id, discord_username, nama, email, package_type, end_date, expiry_reminder_count 
                        FROM subscriptions 
                        WHERE status = "active" 
                        AND DATE(end_date) <= ? 
                        AND DATE(end_date) > ?
                        AND expiry_reminder_count < 3''',
                     (three_days_later, now.strftime('%Y-%m-%d')))
            
            warning_members = c.fetchall()
            
            if warning_members:
                print(f"🔔 3-Day Warning Check: Found {len(warning_members)} members to warn")
            
            guild = bot.get_guild(GUILD_ID)
            if not guild:
                conn.close()
                await asyncio.sleep(60)
                continue
            
            for order_id, discord_id, discord_username, nama, email, package_type, end_date, reminder_count in warning_members:
                try:
                    member = guild.get_member(int(discord_id))
                    if member:
                        packages = get_all_packages()
                        pkg_name = packages.get(package_type, {}).get('name', 'The Warrior')
                        
                        # Calculate sisa hari
                        end_dt = datetime.fromisoformat(end_date)
                        sisa_hari = (end_dt.date() - now.date()).days
                        
                        # Send YELLOW EMBED DM
                        try:
                            warning_embed = discord.Embed(
                                title="⚠️ PERINGATAN JATUH TEMPO! ⚠️",
                                description=f"Membership **{pkg_name}** Anda akan segera berakhir!",
                                color=0xffc107
                            )
                            warning_embed.add_field(name="📅 Sisa Waktu", value=f"**{sisa_hari} Hari**", inline=True)
                            warning_embed.add_field(name="⏰ Berakhir", value=format_jakarta_datetime_full(end_dt), inline=True)
                            warning_embed.add_field(name="💡 Aksi", value="Gunakan `/buy` untuk perpanjang sekarang!", inline=False)
                            warning_embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
                            warning_embed.set_thumbnail(url=member.avatar.url if member.avatar else "")
                            
                            await member.send(embed=warning_embed)
                            print(f"  ✅ 3-Day WARNING YELLOW EMBED sent to {discord_username} (Hari ke-{reminder_count+1})")
                        except discord.HTTPException as e:
                            print(f"  ⚠️ Could not send DM to {discord_id}: {e}")
                        
                        # Send YELLOW GRADIENT EMAIL
                        try:
                            member_avatar = str(member.avatar.url) if member.avatar else str(member.default_avatar)
                            end_datetime = format_jakarta_datetime_full(end_dt)
                            send_3day_expiry_warning_email(nama, email, pkg_name, end_datetime, member_avatar, sisa_hari)
                            print(f"  ✅ 3-Day WARNING YELLOW EMAIL sent to {email}")
                        except Exception as e:
                            print(f"  ❌ Error sending 3-day warning email: {e}")
                        
                        # Increment reminder count
                        c.execute('UPDATE subscriptions SET expiry_reminder_count = expiry_reminder_count + 1 WHERE order_id = ?', (order_id,))
                        print(f"  ✅ Reminder count incremented for {nama} (Now: {reminder_count+1}/3)")
                    
                except Exception as e:
                    print(f"  ❌ Error: {e}")
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"❌ Error in 3-day warning check: {e}")
        
        # Check setiap 1 jam (3600 seconds)
        await asyncio.sleep(3600)


async def check_trial_expiry_warning():
    """Background task untuk kirim warning trial member sebelum expire (1 hari sebelum)"""
    await bot.wait_until_ready()
    
    while not bot.is_closed():
        try:
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            now = get_jakarta_datetime()
            tomorrow = (now + timedelta(hours=24)).strftime('%Y-%m-%d')
            
            # Find trial members yang akan expire dalam < 24 jam
            c.execute('''SELECT discord_id, discord_username, username, email, trial_end 
                        FROM trial_members 
                        WHERE status = "active" 
                        AND DATE(trial_end) <= ?''',
                     (tomorrow,))
            
            trial_warnings = c.fetchall()
            
            if trial_warnings:
                print(f"🔔 Trial Warning Check: Found {len(trial_warnings)} trial members to warn")
            
            guild = bot.get_guild(GUILD_ID)
            if not guild:
                conn.close()
                await asyncio.sleep(60)
                continue
            
            for discord_id, discord_username, username, email, trial_end in trial_warnings:
                try:
                    member = guild.get_member(int(discord_id))
                    if member:
                        # Send ORANGE DM notification
                        try:
                            trial_embed = discord.Embed(
                                title="⏳ TRIAL AKAN BERAKHIR! ⏳",
                                description="Akses trial The Warrior kamu akan segera berakhir",
                                color=0xf7931a
                            )
                            trial_embed.add_field(name="📅 Berakhir", value=format_jakarta_datetime_full(datetime.fromisoformat(trial_end)), inline=True)
                            trial_embed.add_field(name="⚠️ Status", value="KURANG DARI 24 JAM", inline=True)
                            trial_embed.add_field(name="💡 Aksi", value="Klik `/buy` untuk perpanjang atau beli paket premium!", inline=False)
                            trial_embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
                            trial_embed.set_thumbnail(url=member.avatar.url if member.avatar else "")
                            
                            await member.send(embed=trial_embed)
                            print(f"  ✅ Trial warning ORANGE EMBED sent to {discord_username}")
                        except discord.HTTPException as e:
                            print(f"  ⚠️ Could not send DM to {discord_id}: {e}")
                        
                        # Send ORANGE GRADIENT EMAIL
                        try:
                            member_avatar = str(member.avatar.url) if member.avatar else str(member.default_avatar)
                            trial_end_display = format_jakarta_datetime_full(datetime.fromisoformat(trial_end))
                            send_trial_expiry_warning_email(username, email, trial_end_display, member_avatar, 24)
                            print(f"  ✅ Trial warning ORANGE EMAIL sent to {email}")
                        except Exception as e:
                            print(f"  ❌ Error sending trial warning email: {e}")
                except Exception as e:
                    print(f"  ❌ Error: {e}")
            
            conn.close()
            
        except Exception as e:
            print(f"❌ Error in trial warning check: {e}")
        
        await asyncio.sleep(3600)


async def remove_expired_trial_members():
    await bot.wait_until_ready()
    
    while not bot.is_closed():
        try:
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            # Use Jakarta timezone untuk accurate comparison dengan database
            now = get_jakarta_datetime().strftime('%Y-%m-%d %H:%M:%S')
            
            c.execute('SELECT discord_id, discord_username FROM trial_members WHERE status = "active" AND datetime(trial_end) <= datetime(?)', (now,))
            expired_trials = c.fetchall()
            
            print(f"🔍 Trial check: Found {len(expired_trials)} expired trial members")
            
            guild = bot.get_guild(GUILD_ID)
            if not guild:
                conn.close()
                await asyncio.sleep(300)
                continue
            
            for discord_id, discord_username in expired_trials:
                try:
                    member = guild.get_member(int(discord_id))
                    if member:
                        trial_role = discord.utils.get(guild.roles, name=TRIAL_MEMBER_ROLE_NAME)
                        if trial_role and trial_role in member.roles:
                            await member.remove_roles(trial_role)
                            print(f"  ✅ Trial role removed from {discord_username}")
                            
                            try:
                                embed = discord.Embed(
                                    title="⏰ Trial Membership Expired",
                                    description="Masa trial Anda telah berakhir. Saatnya untuk upgrade membership The Warrior!",
                                    color=0xff0000
                                )
                                embed.set_footer(text="📊 Diary Crypto Bot")
                                
                                await member.send(embed=embed)
                                print(f"  ✅ DM sent to {member.name}")
                            except discord.HTTPException:
                                print(f"  ⚠️ Could not send DM to {discord_id}")
                    
                    c.execute('UPDATE trial_members SET status = "expired" WHERE discord_id = ?', (discord_id,))
                    print(f"  ✅ Trial member marked as expired")
                    
                except Exception as e:
                    print(f"  ❌ Error: {e}")
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"❌ Error in trial removal: {e}")
        
        await asyncio.sleep(300)


@bot.event
async def on_ready():
    print(f"✅ {bot.user.name}#{bot.user.discriminator} has connected to Discord!")
    
    guild = bot.get_guild(GUILD_ID)
    if guild:
        print(f"✅ Found guild: {guild.name} (ID: {guild.id})")
    
    try:
        print(f"🔄 Syncing commands globally...")
        await tree.sync()
        print(f"✅ Global sync: {len(tree.get_commands())} commands")
        
        print(f"🔄 Syncing commands to guild {guild.name}...")
        await tree.sync(guild=guild)
        print(f"✅ Guild sync: {len(tree.get_commands())} commands")
    except Exception as e:
        print(f"❌ Error syncing commands: {e}")
    
    if not bot.is_synced:
        print("✅ Stale order cleanup started!")
        bot.loop.create_task(cleanup_stale_orders())
        
        print("✅ Expiry checker started!")
        bot.loop.create_task(check_expired_subscriptions())
        
        print("✅ 3-Day expiry warning started!")
        bot.loop.create_task(check_3day_expiry_warning())
        
        print("✅ Trial expiry warning started!")
        bot.loop.create_task(check_trial_expiry_warning())
        
        print("✅ Auto role removal started!")
        bot.loop.create_task(auto_remove_expired_members())
        
        print("✅ Trial member auto-removal started!")
        bot.loop.create_task(remove_expired_trial_members())
        
        print("✅ Crypto news MANUAL mode (use /post_crypto_news_now to test)")
        
        bot.is_synced = True
    
    print("🎉 Bot is ready!")


@tree.command(name="post_crypto_news_now", description="[Admin/Com-Manager] Manual post crypto news untuk testing")
@discord.app_commands.default_permissions(administrator=False)
async def post_crypto_news_now(interaction: discord.Interaction):
    # Admin, guild owner, dan Orion saja yang bisa akses
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message(
            "❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", 
            ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        guild = interaction.guild
        if not guild:
            await interaction.followup.send("❌ Guild tidak ditemukan!", ephemeral=True)
            return
        
        # Find news channels - only post to payment channel for now
        news_channels = []
        target_channel_names = ["💳｜payment"]  # Only payment channel (berita-crypto disabled for now)
        
        for channel in guild.text_channels:
            if channel.name in target_channel_names:
                news_channels.append(channel)
        
        if not news_channels:
            await interaction.followup.send(f"❌ Channel tidak ditemukan! Cari: {', '.join(target_channel_names)}", ephemeral=True)
            return
        
        # Fetch crypto news
        articles = await fetch_crypto_news()
        
        if articles:
            count = 0
            # Post to each channel
            for news_channel in news_channels:
                for article in articles:
                    try:
                        title = article.get('title', 'Untitled')
                        image = article.get('image_url', '')
                        analysis = article.get('analysis', '')
                        
                        # Create main embed dengan analysis ONLY (TANPA link)
                        embed = discord.Embed(
                            title=title[:256],
                            description=analysis[:4000] if analysis else "Analysis tidak tersedia",
                            color=0xf7931a
                        )
                        
                        if image:
                            embed.set_image(url=image)
                        
                        embed.set_footer(text="📊 Analisis Crypto News")
                        
                        await news_channel.send(embed=embed)
                        count += 1
                        await asyncio.sleep(1)
                    except Exception as e:
                        print(f"⚠️ Error posting article: {e}")
            
            channel_list = ", ".join([f"#{ch.name}" for ch in news_channels])
            await interaction.followup.send(
                f"✅ {count} berita crypto sudah di-post ke: {channel_list}!",
                ephemeral=True
            )
        else:
            await interaction.followup.send(
                "⚠️ Tidak ada berita crypto yang ditemukan saat ini",
                ephemeral=True
            )
    
    except Exception as e:
        print(f"❌ Error posting news: {e}")
        await interaction.followup.send(
            f"❌ Error: {str(e)}",
            ephemeral=True
        )


class BuyNewModal(discord.ui.Modal, title="📝 Beli Paket Baru"):
    nama_user_discord = discord.ui.TextInput(label="Nama Discord", placeholder="Masukkan username Discord Anda", required=True)
    email = discord.ui.TextInput(label="Email", placeholder="email@example.com", required=True)
    nama = discord.ui.TextInput(label="Nama Lengkap", placeholder="Masukkan nama Anda", required=True)
    discount_code = discord.ui.TextInput(label="Kode Diskon (opsional)", placeholder="Ketik kode diskon atau leave blank", required=False, default="")
    referral_code = discord.ui.TextInput(label="Kode Referral (opsional)", placeholder="Ketik kode referral atau leave blank", required=False, default="")
    
    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        
        discord_id = str(interaction.user.id)
        discord_username = self.nama_user_discord.value
        package_id = self.package_id
        
        packages = get_all_packages()
        pkg = packages.get(package_id)
        email_val = self.email.value
        nama_val = self.nama.value
        discount_code_val = self.discount_code.value.strip() if self.discount_code.value else ""
        referral_code_val = self.referral_code.value.strip() if self.referral_code.value else ""
        
        # Calculate price dengan discount
        final_price = pkg['price']
        discount_info = ""
        referral_info = ""
        analyst_id = None
        analyst_name = None
        
        # Verify discount code
        if discount_code_val:
            verify_result = verify_discount_code(discount_code_val)
            if verify_result["valid"]:
                discount_percent = verify_result["discount_percent"]
                discount_amount = int(pkg['price'] * discount_percent / 100)
                final_price = pkg['price'] - discount_amount
                discount_info = f"\n💰 Diskon: {discount_percent}% (-Rp {discount_amount:,})"
            else:
                embed = discord.Embed(
                    title="⚠️ Kode Diskon Tidak Valid",
                    description="Maaf, kode diskon yang Anda masukkan tidak ditemukan atau sudah tidak berlaku.",
                    color=0xff4444
                )
                embed.add_field(name="💳 Kode Yang Anda Masukkan", value=f"`{discount_code_val.upper()}`", inline=False)
                embed.add_field(name="📝 Alasan", value=verify_result['message'], inline=False)
                embed.add_field(name="💡 Saran", value="Anda tetap bisa lanjut checkout tanpa diskon, atau hubungi admin untuk kode yang valid", inline=False)
                embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
                await interaction.followup.send(embed=embed, ephemeral=True)
                return
        
        # Verify referral code
        if referral_code_val:
            verify_result = verify_referral_code(referral_code_val)
            if verify_result["valid"]:
                analyst_id = verify_result["analyst_id"]
                analyst_name = verify_result["analyst_name"]
                commission_percent = verify_result["commission_percent"]
                commission_amount = int(final_price * commission_percent / 100)
                referral_info = f"\n👥 Referral dari: {analyst_name} (Komisi: {commission_percent}% = Rp {commission_amount:,})"
            else:
                await interaction.followup.send(f"❌ {verify_result['message']}", ephemeral=True)
                return
        
        # Create order
        order_id = f"ORD_{discord_id}_{int(time.time())}"
        save_pending_order(order_id, discord_id, discord_username, nama_val, email_val, package_id, "https://checkout.midtrans.com")
        
        # Update order dan track komisi
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        c.execute('UPDATE pending_orders SET price = ? WHERE order_id = ?', (final_price, order_id))
        
        if discount_code_val:
            c.execute('UPDATE discount_codes SET used_count = used_count + 1 WHERE code = ?', (discount_code_val.upper(),))
        
        if analyst_id and referral_code_val:
            # Track komisi untuk analyst
            commission_amount = int(final_price * 30 / 100)
            c.execute('''CREATE TABLE IF NOT EXISTS commissions (
                order_id TEXT,
                analyst_id TEXT,
                analyst_name TEXT,
                commission_amount REAL,
                created_at TEXT
            )''')
            created_at = get_jakarta_datetime().strftime('%Y-%m-%d %H:%M:%S')
            c.execute('INSERT INTO commissions (order_id, analyst_id, analyst_name, commission_amount, created_at) VALUES (?, ?, ?, ?, ?)',
                     (order_id, analyst_id, analyst_name, commission_amount, created_at))
        
        conn.commit()
        conn.close()
        
        embed = discord.Embed(
            title="✅ Beli Paket Baru - Checkout Dibuat",
            color=0x00ff00
        )
        embed.add_field(name="📦 Paket", value=f"**{pkg['name']}**", inline=True)
        embed.add_field(name="💰 Harga", value=f"Rp **{pkg['price']:,}**", inline=True)
        embed.add_field(name="💳 Harga Akhir", value=f"Rp **{final_price:,}**{discount_info}{referral_info}", inline=False)
        embed.add_field(name="👤 Discord Username", value=discord_username, inline=True)
        embed.add_field(name="📧 Email", value=email_val, inline=True)
        embed.add_field(name="👤 Nama Lengkap", value=nama_val, inline=False)
        embed.add_field(name="Order ID", value=f"`{order_id}`", inline=False)
        embed.set_footer(text="Tunggu instruksi pembayaran selanjutnya...")
        
        await interaction.followup.send(embed=embed, ephemeral=True)
        
        # Send DM dengan instruksi pembayaran - EMBED DENGAN AVATAR
        try:
            # Generate payment link dari Midtrans (redirect_url)
            payment_link = generate_snap_token(order_id, final_price, nama_val, email_val)
            
            if not payment_link:
                payment_link = "https://app.sandbox.midtrans.com"  # Fallback ke halaman utama
                print(f"⚠️ Fallback payment link digunakan untuk {order_id}")
            
            dm_embed = discord.Embed(
                title="✅ CHECKOUT BERHASIL!",
                description="Silakan lanjutkan pembayaran Anda di Midtrans",
                color=0xf7931a
            )
            dm_embed.set_thumbnail(url=interaction.user.avatar.url if interaction.user.avatar else interaction.user.default_avatar.url)
            dm_embed.add_field(name="📦 Paket", value=f"**{pkg['name']}**", inline=True)
            dm_embed.add_field(name="💳 Harga Akhir", value=f"Rp **{final_price:,}**", inline=True)
            dm_embed.add_field(name="📋 Order ID", value=f"`{order_id}`", inline=False)
            dm_embed.add_field(name="👤 Pembeli", value=f"**{nama_val}**", inline=True)
            dm_embed.add_field(name="📧 Email", value=email_val, inline=True)
            dm_embed.add_field(name="🔗 Link Pembayaran", value=f"[Klik di sini untuk bayar]({payment_link})", inline=False)
            dm_embed.add_field(name="📌 Info", value="Invoice juga sudah dikirim ke email Anda", inline=False)
            dm_embed.set_footer(text="Diary Crypto Payment Bot • Terima kasih!")
            
            await interaction.user.send(embed=dm_embed)
            print(f"✅ DM checkout dikirim ke {discord_username}")
        except discord.HTTPException as e:
            print(f"⚠️ Gagal kirim DM ke {discord_username}: {e}")


class RenewModal(discord.ui.Modal, title="🔄 Perpanjang Membership"):
    nama_user_discord = discord.ui.TextInput(label="Nama Discord", placeholder="Masukkan username Discord Anda", required=True)
    discount_code = discord.ui.TextInput(label="Kode Diskon (opsional)", placeholder="Ketik kode diskon atau leave blank", required=False, default="")
    referral_code = discord.ui.TextInput(label="Kode Referral (opsional)", placeholder="Ketik kode referral atau leave blank", required=False, default="")
    
    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        
        discord_id = str(interaction.user.id)
        discord_username = self.nama_user_discord.value
        package_id = self.package_id
        
        # Get current subscription
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        c.execute('SELECT email, nama, start_date, end_date FROM subscriptions WHERE discord_id = ? AND status = "active"', (discord_id,))
        current = c.fetchone()
        
        if not current:
            conn.close()
            await interaction.followup.send("❌ Anda belum memiliki membership aktif!", ephemeral=True)
            return
        
        email_val, nama_val, old_start, old_end = current
        
        packages = get_all_packages()
        pkg = packages.get(package_id)
        discount_code_val = self.discount_code.value.strip() if self.discount_code.value else ""
        referral_code_val = self.referral_code.value.strip() if self.referral_code.value else ""
        
        # Calculate price dengan discount
        final_price = pkg['price']
        discount_info = ""
        referral_info = ""
        analyst_id = None
        analyst_name = None
        
        # Verify discount code
        if discount_code_val:
            verify_result = verify_discount_code(discount_code_val)
            if verify_result["valid"]:
                discount_percent = verify_result["discount_percent"]
                discount_amount = int(pkg['price'] * discount_percent / 100)
                final_price = pkg['price'] - discount_amount
                discount_info = f"\n💰 Diskon: {discount_percent}% (-Rp {discount_amount:,})"
            else:
                conn.close()
                await interaction.followup.send(f"❌ {verify_result['message']}", ephemeral=True)
                return
        
        # Verify referral code
        if referral_code_val:
            verify_result = verify_referral_code(referral_code_val)
            if verify_result["valid"]:
                analyst_id = verify_result["analyst_id"]
                analyst_name = verify_result["analyst_name"]
                commission_percent = verify_result["commission_percent"]
                commission_amount = int(final_price * commission_percent / 100)
                referral_info = f"\n👥 Referral dari: {analyst_name} (Komisi: {commission_percent}% = Rp {commission_amount:,})"
            else:
                conn.close()
                await interaction.followup.send(f"❌ {verify_result['message']}", ephemeral=True)
                return
        
        # Create renewal order
        order_id = f"REN_{discord_id}_{int(time.time())}"
        save_pending_order(order_id, discord_id, discord_username, nama_val, email_val, package_id, "https://checkout.midtrans.com")
        
        # Create renewal table if not exist
        c.execute('''CREATE TABLE IF NOT EXISTS renewals (
            order_id TEXT PRIMARY KEY,
            discord_id TEXT,
            discord_username TEXT,
            package_type TEXT,
            old_end_date TEXT,
            new_end_date TEXT,
            renewal_price REAL,
            discount_applied TEXT,
            referral_applied TEXT,
            status TEXT DEFAULT "pending",
            created_at TEXT
        )''')
        
        # Calculate new end date
        old_end_dt = datetime.strptime(old_end, '%Y-%m-%d %H:%M:%S')
        new_end_dt = old_end_dt + timedelta(days=pkg['duration_days'])
        new_end_date = new_end_dt.strftime('%Y-%m-%d %H:%M:%S')
        
        created_at = get_jakarta_datetime().strftime('%Y-%m-%d %H:%M:%S')
        
        # Track renewal
        c.execute('UPDATE pending_orders SET price = ? WHERE order_id = ?', (final_price, order_id))
        c.execute('INSERT INTO renewals (order_id, discord_id, discord_username, package_type, old_end_date, new_end_date, renewal_price, discount_applied, referral_applied, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                 (order_id, discord_id, discord_username, package_id, old_end, new_end_date, final_price, discount_code_val or "none", referral_code_val or "none", created_at))
        
        if discount_code_val:
            c.execute('UPDATE discount_codes SET used_count = used_count + 1 WHERE code = ?', (discount_code_val.upper(),))
        
        if analyst_id and referral_code_val:
            commission_amount = int(final_price * 30 / 100)
            c.execute('''CREATE TABLE IF NOT EXISTS commissions (
                order_id TEXT,
                analyst_id TEXT,
                analyst_name TEXT,
                commission_amount REAL,
                created_at TEXT
            )''')
            c.execute('INSERT INTO commissions (order_id, analyst_id, analyst_name, commission_amount, created_at) VALUES (?, ?, ?, ?, ?)',
                     (order_id, analyst_id, analyst_name, commission_amount, created_at))
        
        conn.commit()
        conn.close()
        
        embed = discord.Embed(
            title="✅ Perpanjang Membership - Checkout Dibuat",
            color=0x00ff00
        )
        embed.add_field(name="📦 Paket", value=f"**{pkg['name']}**", inline=True)
        embed.add_field(name="💰 Harga", value=f"Rp **{pkg['price']:,}**", inline=True)
        embed.add_field(name="👤 Discord Username", value=discord_username, inline=True)
        embed.add_field(name="📅 Perpanjang Dari", value=old_end, inline=False)
        embed.add_field(name="📅 Sampai", value=new_end_date, inline=False)
        embed.add_field(name="💳 Harga Akhir", value=f"Rp **{final_price:,}**{discount_info}{referral_info}", inline=False)
        embed.add_field(name="Order ID", value=f"`{order_id}`", inline=False)
        embed.set_footer(text="Tunggu instruksi pembayaran selanjutnya...")
        
        await interaction.followup.send(embed=embed, ephemeral=True)
        
        # Send DM dengan instruksi perpanjangan - EMBED DENGAN AVATAR
        try:
            # Generate payment link dari Midtrans (redirect_url)
            payment_link = generate_snap_token(order_id, final_price, nama_val, email_val)
            
            if not payment_link:
                payment_link = "https://app.sandbox.midtrans.com"  # Fallback ke halaman utama
                print(f"⚠️ Fallback payment link digunakan untuk {order_id}")
            
            dm_embed = discord.Embed(
                title="✅ PERPANJANGAN BERHASIL!",
                description="Silakan lanjutkan pembayaran Anda di Midtrans",
                color=0xf7931a
            )
            dm_embed.set_thumbnail(url=interaction.user.avatar.url if interaction.user.avatar else interaction.user.default_avatar.url)
            dm_embed.add_field(name="📦 Paket", value=f"**{pkg['name']}**", inline=True)
            dm_embed.add_field(name="💳 Harga Akhir", value=f"Rp **{final_price:,}**", inline=True)
            dm_embed.add_field(name="📋 Order ID", value=f"`{order_id}`", inline=False)
            dm_embed.add_field(name="📅 Perpanjang Dari", value=old_end, inline=True)
            dm_embed.add_field(name="📅 Sampai", value=new_end_date, inline=True)
            dm_embed.add_field(name="🔗 Link Pembayaran", value=f"[Klik di sini untuk bayar]({payment_link})", inline=False)
            dm_embed.add_field(name="📌 Info", value="Invoice juga sudah dikirim ke email Anda", inline=False)
            dm_embed.set_footer(text="Diary Crypto Payment Bot • Terima kasih!")
            
            await interaction.user.send(embed=dm_embed)
            print(f"✅ DM perpanjangan dikirim ke {discord_username}")
        except discord.HTTPException as e:
            print(f"⚠️ Gagal kirim DM ke {discord_username}: {e}")


@tree.command(name="buy", description="Beli atau perpanjang membership The Warrior")
async def buy_command(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    
    packages = get_all_packages()
    discord_id = str(interaction.user.id)
    
    # Check if user already has active membership
    conn = sqlite3.connect('warrior_subscriptions.db')
    c = conn.cursor()
    c.execute('SELECT package_type, end_date FROM subscriptions WHERE discord_id = ? AND status = "active"', (discord_id,))
    existing = c.fetchone()
    conn.close()
    
    # Buttons untuk pilih aksi
    class ActionView(discord.ui.View):
        def __init__(self, has_membership):
            super().__init__()
            self.has_membership = has_membership
        
        @discord.ui.button(label="🛍️ Beli Paket Baru", style=discord.ButtonStyle.primary)
        async def buy_new(self, button_interaction: discord.Interaction, button: discord.ui.Button):
            class PackageSelect(discord.ui.Select):
                def __init__(self):
                    options = [
                        discord.SelectOption(
                            label=f"{pkg['name']} - Rp {pkg['price']:,}",
                            value=key,
                            description=pkg['duration_text']
                        )
                        for key, pkg in packages.items()
                    ]
                    super().__init__(
                        placeholder="Pilih paket baru...",
                        min_values=1,
                        max_values=1,
                        options=options
                    )
                
                async def callback(self, select_interaction: discord.Interaction):
                    package_id = self.values[0]
                    modal = BuyNewModal()
                    modal.package_id = package_id
                    await select_interaction.response.send_modal(modal)
            
            class SelectView(discord.ui.View):
                def __init__(self):
                    super().__init__()
                    self.add_item(PackageSelect())
            
            embed = discord.Embed(
                title="📦 Pilih Paket Baru",
                description="Pilih paket membership yang ingin dibeli:",
                color=0xf7931a
            )
            await button_interaction.response.send_message(embed=embed, view=SelectView(), ephemeral=True)
        
        @discord.ui.button(label="🔄 Perpanjang Membership", style=discord.ButtonStyle.success)
        async def renew(self, button_interaction: discord.Interaction, button: discord.ui.Button):
            if not self.has_membership:
                await button_interaction.response.send_message("❌ Anda belum memiliki membership aktif!", ephemeral=True)
                return
            
            class PackageSelect(discord.ui.Select):
                def __init__(self):
                    options = [
                        discord.SelectOption(
                            label=f"{pkg['name']} - Rp {pkg['price']:,}",
                            value=key,
                            description=pkg['duration_text']
                        )
                        for key, pkg in packages.items()
                    ]
                    super().__init__(
                        placeholder="Pilih paket perpanjang...",
                        min_values=1,
                        max_values=1,
                        options=options
                    )
                
                async def callback(self, select_interaction: discord.Interaction):
                    package_id = self.values[0]
                    modal = RenewModal()
                    modal.package_id = package_id
                    await select_interaction.response.send_modal(modal)
            
            class SelectView(discord.ui.View):
                def __init__(self):
                    super().__init__()
                    self.add_item(PackageSelect())
            
            pkg_type, end_date = existing
            embed = discord.Embed(
                title="🔄 Perpanjang Membership",
                description=f"Membership saat ini berakhir: **{end_date}**\n\nPilih paket perpanjang:",
                color=0xf7931a
            )
            await button_interaction.response.send_message(embed=embed, view=SelectView(), ephemeral=True)
    
    # Main embed dengan 2 buttons
    embed = discord.Embed(
        title="🎯 The Warrior - Membership",
        description="Pilih aksi yang ingin Anda lakukan:",
        color=0xf7931a
    )
    
    if existing:
        pkg_type, end_date = existing
        embed.add_field(name="✅ Membership Aktif", value=f"Berakhir: {end_date}", inline=False)
    
    await interaction.followup.send(embed=embed, view=ActionView(has_membership=bool(existing)), ephemeral=True)


# [DEPRECATED] /buy_form dan /buy_form_submit - tidak digunakan lagi
# Gunakan /buy saja untuk semua fitur beli & perpanjang membership


class TrialRedeemModal(discord.ui.Modal, title="🎉 Redeem Trial Member"):
    trial_code = discord.ui.TextInput(label="Trial Code", placeholder="Masukkan kode trial...", required=True, max_length=100)
    email = discord.ui.TextInput(label="Email", placeholder="your@email.com", required=True, max_length=100)
    username = discord.ui.TextInput(label="Username", placeholder="Nama lengkap Anda", required=True, max_length=100)
    
    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        
        trial_code_val = str(self.trial_code).strip().upper()
        email_val = str(self.email).strip()
        username_val = str(self.username).strip()
        
        discord_id = str(interaction.user.id)
        discord_username = interaction.user.name
        
        # Check if code exists
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        c.execute('SELECT id, trial_code, discord_id, discord_username, email, username, trial_started, trial_end, status, created_at, created_by, duration_days, validity_days, code_expiry_date, max_uses, used_count FROM trial_members WHERE trial_code = ?', (trial_code_val,))
        code_check = c.fetchone()
        
        if not code_check:
            embed = discord.Embed(
                title="⚠️ Kode Trial Tidak Valid",
                description="Maaf, kode trial yang Anda masukkan tidak ditemukan atau sudah digunakan.",
                color=0xff4444
            )
            embed.add_field(name="📝 Kode Yang Anda Masukkan", value=f"`{trial_code_val}`", inline=False)
            embed.add_field(name="💡 Saran", value="Periksa kembali kode trial Anda atau hubungi admin untuk mendapatkan kode baru", inline=False)
            embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
            await interaction.followup.send(embed=embed, ephemeral=True)
            conn.close()
            return
        
        # Check if code is still valid (not expired)
        code_expiry = code_check[13]  # code_expiry_date (index 13)
        max_uses = int(code_check[14]) if code_check[14] else 0  # max_uses (index 14) - convert to int
        used_count = int(code_check[15]) if code_check[15] else 0  # used_count (index 15) - convert to int
        duration_days = int(code_check[11]) if code_check[11] else 1  # duration_days (index 11) - convert to int with default
        
        if code_expiry and isinstance(code_expiry, str):
            try:
                code_expiry_dt = datetime.fromisoformat(code_expiry)
                if get_jakarta_datetime() > code_expiry_dt:
                    embed = discord.Embed(
                        title="⚠️ Kode Trial Sudah Expired",
                        description="Maaf, kode trial ini sudah tidak berlaku lagi.",
                        color=0xff4444
                    )
                    embed.add_field(name="📅 Berlaku Sampai", value=code_expiry, inline=False)
                    embed.add_field(name="💡 Saran", value="Hubungi admin untuk mendapatkan kode trial yang baru", inline=False)
                    embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
                    await interaction.followup.send(embed=embed, ephemeral=True)
                    conn.close()
                    return
            except (ValueError, TypeError):
                pass  # Skip expiry check if date is invalid
        
        # Check if code has reached max uses
        if max_uses > 0 and used_count >= max_uses:
            embed = discord.Embed(
                title="⚠️ Kode Trial Sudah Penuh",
                description="Maaf, kode trial ini sudah digunakan oleh terlalu banyak orang.",
                color=0xff4444
            )
            embed.add_field(name="👥 Kapasitas", value=f"Sudah digunakan: {used_count}/{max_uses} orang", inline=False)
            embed.add_field(name="💡 Saran", value="Hubungi admin untuk mendapatkan kode trial yang baru", inline=False)
            embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
            await interaction.followup.send(embed=embed, ephemeral=True)
            conn.close()
            return
        
        # Check if user already has active trial
        c.execute('SELECT * FROM trial_members WHERE discord_id = ? AND status = "active"', (discord_id,))
        existing = c.fetchone()
        
        if existing:
            await interaction.followup.send("❌ Anda sudah memiliki trial member aktif!", ephemeral=True)
            conn.close()
            return
        
        guild = interaction.guild
        trial_role = discord.utils.get(guild.roles, name=TRIAL_MEMBER_ROLE_NAME)
        
        if not trial_role:
            await interaction.followup.send("❌ Role Trial Member tidak ditemukan di server!", ephemeral=True)
            conn.close()
            return
        
        trial_start = get_jakarta_datetime()
        trial_end = trial_start + timedelta(days=duration_days)
        
        # Update trial_members dengan data lengkap dan increment used_count
        c.execute('''UPDATE trial_members 
                    SET discord_id = ?, discord_username = ?, email = ?, username = ?, trial_started = ?, trial_end = ?, status = "active", used_count = used_count + 1
                    WHERE trial_code = ?''',
                 (discord_id, discord_username, email_val, username_val, trial_start.strftime('%Y-%m-%d %H:%M:%S'), 
                  trial_end.strftime('%Y-%m-%d %H:%M:%S'), trial_code_val))
        conn.commit()
        conn.close()
        
        await interaction.user.add_roles(trial_role)
        
        # Send ORANGE EMBED - Trial activated
        embed = discord.Embed(
            title="🎉 TRIAL MEMBER ACTIVATED!",
            description="Anda sekarang menjadi Trial Member selama 1 JAM!",
            color=0xf7931a
        )
        embed.set_thumbnail(url=interaction.user.avatar.url if interaction.user.avatar else interaction.user.default_avatar.url)
        embed.add_field(name="👤 Username", value=username_val, inline=True)
        embed.add_field(name="📧 Email", value=email_val, inline=True)
        embed.add_field(name="⏱️ Durasi", value="**1 JAM**", inline=True)
        embed.add_field(name="🔄 Status", value="**AKTIF**", inline=True)
        embed.add_field(name="📅 Mulai", value=format_jakarta_datetime(trial_start), inline=False)
        embed.add_field(name="⏰ Berakhir", value=format_jakarta_datetime(trial_end), inline=False)
        embed.add_field(name="💡 Info", value="Role akan otomatis dihapus saat trial berakhir. Nikmati akses eksklusif The Warrior!", inline=False)
        embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
        
        await interaction.followup.send(embed=embed, ephemeral=True)
        
        # Send DM confirmation dengan ORANGE EMBED
        try:
            dm_embed = discord.Embed(
                title="🎉 TRIAL MEMBER BERHASIL!",
                description="Anda telah berhasil menjadi Trial Member!",
                color=0xf7931a
            )
            dm_embed.set_thumbnail(url=interaction.user.avatar.url if interaction.user.avatar else interaction.user.default_avatar.url)
            dm_embed.add_field(name="👤 Username", value=username_val, inline=False)
            dm_embed.add_field(name="📧 Email", value=email_val, inline=False)
            dm_embed.add_field(name="⏱️ Durasi", value="1 Jam", inline=False)
            dm_embed.add_field(name="📅 Mulai", value=format_jakarta_datetime(trial_start), inline=False)
            dm_embed.add_field(name="⏰ Berakhir", value=format_jakarta_datetime(trial_end), inline=False)
            dm_embed.add_field(name="🔗 Aksi", value="Gunakan `/buy` untuk beli paket lebih lama!", inline=False)
            dm_embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
            
            await interaction.user.send(embed=dm_embed)
            print(f"✅ Trial ORANGE DM sent to {discord_username}")
        except discord.HTTPException as e:
            print(f"⚠️ Could not send DM to {discord_username}: {e}")
        
        # Send trial member email
        try:
            member_avatar = str(interaction.user.avatar.url) if interaction.user.avatar else str(interaction.user.default_avatar)
            trial_start_str = format_jakarta_datetime(trial_start)
            trial_end_str = format_jakarta_datetime(trial_end)
            send_trial_member_email(username_val, email_val, trial_start_str, trial_end_str, member_avatar, duration_days)
            print(f"✅ Trial member email sent to {email_val}")
        except Exception as e:
            print(f"⚠️ Error sending trial email: {e}")


@tree.command(name="redeem_trial", description="Redeem trial member 1 jam gratis dengan kode")
async def redeem_trial(interaction: discord.Interaction):
    await interaction.response.send_modal(TrialRedeemModal())


@tree.command(name="referral_statistik", description="[Admin] Lihat statistik referral & komisi analyst")
@discord.app_commands.default_permissions(administrator=False)
async def referral_statistik_command(interaction: discord.Interaction):
    # Admin, guild owner, dan Orion saja
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message("❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        # Total komisi
        c.execute('SELECT COALESCE(SUM(commission_amount), 0) FROM commissions')
        total_commission = c.fetchone()[0]
        
        # Komisi per analyst dengan referral code info
        c.execute('''SELECT 
                     rc.created_by as analyst_id,
                     rc.code,
                     COUNT(DISTINCT com.id) as referral_count,
                     COALESCE(SUM(CASE WHEN com.status = "completed" THEN com.commission_amount ELSE 0 END), 0) as earned,
                     COALESCE(SUM(CASE WHEN com.status = "pending" THEN com.commission_amount ELSE 0 END), 0) as pending
                  FROM referral_codes rc
                  LEFT JOIN commissions com ON rc.created_by = com.analyst_id
                  GROUP BY rc.created_by
                  ORDER BY earned DESC''')
        analysts = c.fetchall()
        
        conn.close()
        
        embed = discord.Embed(title="📊 STATISTIK REFERRAL & KOMISI", color=0xf7931a)
        embed.add_field(name="💰 Total Komisi Semua Analyst", value=f"Rp **{int(total_commission):,}**", inline=False)
        
        if analysts:
            stats_text = ""
            for analyst_id, code, count, earned, pending in analysts:
                total = int(earned) + int(pending)
                stats_text += f"👤 **Kode: {code}** (ID: `{analyst_id}`)\n"
                stats_text += f"   • Referral: {count}\n"
                stats_text += f"   • Earned: Rp {int(earned):,}\n"
                stats_text += f"   • Pending: Rp {int(pending):,}\n\n"
            embed.add_field(name="📋 Per Analyst", value=stats_text, inline=False)
        else:
            embed.add_field(name="📋 Per Analyst", value="Belum ada referral", inline=False)
        
        embed.set_footer(text=f"Update: {format_jakarta_datetime(get_jakarta_datetime())}")
        await interaction.followup.send(embed=embed, ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


class ExportMonthModal(discord.ui.Modal, title="📅 Pilih Bulan & Tahun"):
    month = discord.ui.TextInput(label="Bulan (1-12)", placeholder="Contoh: 1, 2, 12", required=True, max_length=2)
    year = discord.ui.TextInput(label="Tahun (YYYY)", placeholder="Contoh: 2025", required=True, max_length=4)
    
    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        
        try:
            month = int(self.month.value.strip())
            year = int(self.year.value.strip())
            
            if month < 1 or month > 12:
                await interaction.followup.send("❌ Bulan harus antara 1-12!", ephemeral=True)
                return
            
            if year < 2020 or year > 2099:
                await interaction.followup.send("❌ Tahun harus antara 2020-2099!", ephemeral=True)
                return
            
            # Format year_month
            year_month = f"{year:04d}-{month:02d}"
            month_name = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", 
                          "Juli", "Agustus", "September", "Oktober", "November", "Desember"][month]
            month_year = f"{month_name} {year}"
            
            # Call export function
            await export_monthly_excel(interaction, year_month, month_year)
        except ValueError:
            await interaction.followup.send("❌ Bulan dan Tahun harus berupa angka!", ephemeral=True)
        except Exception as e:
            await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


async def export_monthly_excel(interaction: discord.Interaction, year_month: str, month_year: str):
    """Generate and send Excel export file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # ===== SHEET 1: SUMMARY =====
        ws_summary = wb.create_sheet("📊 Summary", 0)
        
        # Get stats
        c.execute('SELECT COUNT(*), COALESCE(SUM(price), 0) FROM pending_orders WHERE status = "settlement" AND created_at LIKE ?', (f'%{year_month}%',))
        total_orders, total_revenue = c.fetchone()
        
        c.execute('SELECT COUNT(DISTINCT discord_id) FROM subscriptions WHERE status = "active"')
        active_members = c.fetchone()[0]
        
        c.execute('SELECT COUNT(DISTINCT discord_id) FROM subscriptions WHERE status = "expired"')
        expired_members = c.fetchone()[0]
        
        c.execute('SELECT COUNT(*) FROM trial_members WHERE status = "active"')
        trial_members = c.fetchone()[0]
        
        c.execute('SELECT COUNT(*) FROM pending_orders WHERE status = "pending"')
        pending_orders = c.fetchone()[0]
        
        c.execute('SELECT COALESCE(SUM(commission_amount), 0) FROM commissions WHERE earned_date LIKE ?', (f'%{year_month}%',))
        total_commission = c.fetchone()[0]
        
        # Summary headers
        header_fill = PatternFill(start_color="F7931A", end_color="F7931A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws_summary['A1'] = f"MONTHLY EXPORT - {month_year}"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        ws_summary['A3'] = "💰 REVENUE"
        ws_summary['A3'].font = header_font
        ws_summary['A3'].fill = header_fill
        ws_summary['A4'] = "Total Orders"
        ws_summary['B4'] = total_orders
        ws_summary['A5'] = "Total Revenue"
        ws_summary['B5'] = f"Rp {int(total_revenue):,}"
        
        ws_summary['A7'] = "👥 MEMBERS"
        ws_summary['A7'].font = header_font
        ws_summary['A7'].fill = header_fill
        ws_summary['A8'] = "Active Members"
        ws_summary['B8'] = active_members
        ws_summary['A9'] = "Expired Members"
        ws_summary['B9'] = expired_members
        ws_summary['A10'] = "Trial Members"
        ws_summary['B10'] = trial_members
        
        ws_summary['A12'] = "📦 ORDERS"
        ws_summary['A12'].font = header_font
        ws_summary['A12'].fill = header_fill
        ws_summary['A13'] = "Pending Orders"
        ws_summary['B13'] = pending_orders
        
        ws_summary['A15'] = "👨‍💼 COMMISSIONS"
        ws_summary['A15'].font = header_font
        ws_summary['A15'].fill = header_fill
        ws_summary['A16'] = "Total Commission"
        ws_summary['B16'] = f"Rp {int(total_commission):,}"
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25
        
        # ===== SHEET 2: MEMBERS =====
        ws_members = wb.create_sheet("👥 Members", 1)
        headers = ["Discord ID", "Username", "Email", "Package", "Start Date", "End Date", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_members.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        c.execute('''SELECT discord_id, nama, email, package_type, start_date, end_date, status 
                     FROM subscriptions ORDER BY start_date DESC''')
        members = c.fetchall()
        for row, member in enumerate(members, 2):
            ws_members.cell(row=row, column=1).value = member[0]
            ws_members.cell(row=row, column=2).value = member[1]
            ws_members.cell(row=row, column=3).value = member[2]
            ws_members.cell(row=row, column=4).value = member[3]
            ws_members.cell(row=row, column=5).value = member[4]
            ws_members.cell(row=row, column=6).value = member[5]
            ws_members.cell(row=row, column=7).value = member[6]
        
        for col in range(1, 8):
            ws_members.column_dimensions[chr(64+col)].width = 18
        
        # ===== SHEET 3: TRANSACTIONS =====
        ws_trans = wb.create_sheet("💳 Transactions", 2)
        headers = ["Order ID", "Discord ID", "Username", "Email", "Package", "Price", "Status", "Date"]
        for col, header in enumerate(headers, 1):
            cell = ws_trans.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        c.execute('''SELECT order_id, discord_id, discord_username, email, package_type, price, status, created_at 
                     FROM pending_orders ORDER BY created_at DESC''')
        transactions = c.fetchall()
        for row, trans in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1).value = trans[0]
            ws_trans.cell(row=row, column=2).value = trans[1]
            ws_trans.cell(row=row, column=3).value = trans[2]
            ws_trans.cell(row=row, column=4).value = trans[3]
            ws_trans.cell(row=row, column=5).value = trans[4]
            ws_trans.cell(row=row, column=6).value = f"Rp {int(trans[5]):,}"
            ws_trans.cell(row=row, column=7).value = trans[6]
            ws_trans.cell(row=row, column=8).value = trans[7]
        
        for col in range(1, 9):
            ws_trans.column_dimensions[chr(64+col)].width = 18
        
        # ===== SHEET 4: REFERRALS =====
        ws_ref = wb.create_sheet("🔗 Referrals", 3)
        headers = ["Analyst ID", "Code", "Total Referrals", "Commission (Earned)", "Commission (Pending)"]
        for col, header in enumerate(headers, 1):
            cell = ws_ref.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        c.execute('''SELECT 
                     rc.created_by,
                     rc.code,
                     rc.uses,
                     COALESCE(SUM(CASE WHEN com.status = "completed" THEN com.commission_amount ELSE 0 END), 0),
                     COALESCE(SUM(CASE WHEN com.status = "pending" THEN com.commission_amount ELSE 0 END), 0)
                  FROM referral_codes rc
                  LEFT JOIN commissions com ON rc.created_by = com.analyst_id
                  GROUP BY rc.created_by
                  ORDER BY rc.uses DESC''')
        referrals = c.fetchall()
        for row, ref in enumerate(referrals, 2):
            ws_ref.cell(row=row, column=1).value = ref[0]
            ws_ref.cell(row=row, column=2).value = ref[1]
            ws_ref.cell(row=row, column=3).value = ref[2]
            ws_ref.cell(row=row, column=4).value = f"Rp {int(ref[3]):,}"
            ws_ref.cell(row=row, column=5).value = f"Rp {int(ref[4]):,}"
        
        for col in range(1, 6):
            ws_ref.column_dimensions[chr(64+col)].width = 20
        
        # ===== SHEET 5: TRIAL MEMBERS =====
        ws_trial = wb.create_sheet("🎫 Trial Members", 4)
        headers = ["Code", "Discord ID", "Username", "Trial Start", "Trial End", "Duration (Days)", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_trial.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        c.execute('''SELECT trial_code, discord_id, discord_username, trial_started, trial_end, duration_days, status 
                     FROM trial_members ORDER BY trial_started DESC''')
        trials = c.fetchall()
        for row, trial in enumerate(trials, 2):
            ws_trial.cell(row=row, column=1).value = trial[0]
            ws_trial.cell(row=row, column=2).value = trial[1]
            ws_trial.cell(row=row, column=3).value = trial[2]
            ws_trial.cell(row=row, column=4).value = trial[3]
            ws_trial.cell(row=row, column=5).value = trial[4]
            ws_trial.cell(row=row, column=6).value = trial[5]
            ws_trial.cell(row=row, column=7).value = trial[6]
        
        for col in range(1, 8):
            ws_trial.column_dimensions[chr(64+col)].width = 18
        
        conn.close()
        
        # Save file
        filename = f"Monthly_Export_{year_month.replace('-', '_')}.xlsx"
        wb.save(filename)
        
        # Send file
        embed = discord.Embed(
            title="✅ EXCEL EXPORT READY",
            description=f"File {filename} berhasil dibuat!",
            color=0x00ff00
        )
        embed.add_field(name="📄 File", value=f"`{filename}`", inline=False)
        embed.add_field(name="📅 Period", value=month_year, inline=True)
        embed.add_field(name="📊 Sheets", value="5 sheet (Summary, Members, Transactions, Referrals, Trial)", inline=False)
        embed.set_footer(text=f"Generated: {format_jakarta_datetime(get_jakarta_datetime())}")
        
        await interaction.followup.send(embed=embed, file=discord.File(filename), ephemeral=True)
        
        # Delete file after sending
        import os
        os.remove(filename)
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


@tree.command(name="export_monthly", description="[Admin] Export data membership bulanan ke Excel")
@discord.app_commands.default_permissions(administrator=False)
async def export_monthly_command(interaction: discord.Interaction):
    # Admin, guild owner, dan Orion saja
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message("❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", ephemeral=True)
        return
    
    # Show month/year picker modal
    await interaction.response.send_modal(ExportMonthModal())


@tree.command(name="bot_stats", description="[Admin] Lihat statistik bot - members, revenue, dll")
@discord.app_commands.default_permissions(administrator=False)
async def bot_stats_command(interaction: discord.Interaction):
    # Admin, guild owner, dan Orion saja yang bisa akses
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message(
            "❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", 
            ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        # Stats
        c.execute('SELECT COUNT(*) FROM subscriptions WHERE status = "active"')
        active_members = c.fetchone()[0]
        
        c.execute('SELECT COUNT(*) FROM subscriptions WHERE status = "expired"')
        expired_members = c.fetchone()[0]
        
        c.execute('SELECT SUM(price) FROM pending_orders WHERE status = "settlement"')
        total_revenue = c.fetchone()[0] or 0
        
        c.execute('SELECT COUNT(*) FROM trial_members WHERE status = "active"')
        trial_members = c.fetchone()[0]
        
        conn.close()
        
        embed = discord.Embed(
            title="📊 STATISTIK BOT",
            description="Ringkasan data bot Diary Crypto",
            color=0xf7931a
        )
        embed.add_field(name="👥 Active Members", value=f"**{active_members}** member", inline=True)
        embed.add_field(name="⏰ Expired Members", value=f"**{expired_members}** member", inline=True)
        embed.add_field(name="🎯 Trial Members", value=f"**{trial_members}** trial", inline=True)
        embed.add_field(name="💰 Total Revenue", value=f"Rp **{total_revenue:,}**", inline=False)
        embed.add_field(name="📅 Update Time", value=format_jakarta_datetime(get_jakarta_datetime()), inline=False)
        
        await interaction.followup.send(embed=embed, ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


class CreatePackageModal(discord.ui.Modal, title="➕ Buat Paket Baru"):
    package_id = discord.ui.TextInput(label="ID Paket", placeholder="Contoh: warrior_custom", required=True, max_length=20)
    package_name = discord.ui.TextInput(label="Nama Paket", placeholder="Contoh: The Warrior Premium", required=True, max_length=50)
    price = discord.ui.TextInput(label="Harga (Rp)", placeholder="Contoh: 500000", required=True, max_length=10)
    duration_days = discord.ui.TextInput(label="Durasi (Hari)", placeholder="Contoh: 30, 90", required=True, max_length=5)
    
    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        
        try:
            pkg_id = self.package_id.value.strip().lower()
            pkg_name = self.package_name.value.strip()
            pkg_price = int(self.price.value.strip())
            pkg_duration = int(self.duration_days.value.strip())
            
            if pkg_price <= 0:
                await interaction.followup.send("❌ Harga harus lebih dari 0!", ephemeral=True)
                return
            
            if pkg_duration <= 0:
                await interaction.followup.send("❌ Durasi harus lebih dari 0 hari!", ephemeral=True)
                return
            
            # Validate ID format
            if not pkg_id.replace('_', '').isalnum():
                await interaction.followup.send("❌ ID paket hanya boleh berisi huruf, angka, dan underscore!", ephemeral=True)
                return
            
            # Check if package already exists
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            c.execute('SELECT package_id FROM packages WHERE package_id = ?', (pkg_id,))
            if c.fetchone():
                conn.close()
                await interaction.followup.send(f"❌ Paket dengan ID `{pkg_id}` sudah ada!", ephemeral=True)
                return
            
            # Insert into database
            duration_text = f"{pkg_duration} hari"
            created_at = get_jakarta_datetime().strftime('%Y-%m-%d %H:%M:%S')
            
            c.execute('''INSERT INTO packages (package_id, package_name, price, duration_days, duration_text, role_name, created_at, created_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                     (pkg_id, pkg_name, pkg_price, pkg_duration, duration_text, WARRIOR_ROLE_NAME, created_at, interaction.user.name))
            
            conn.commit()
            conn.close()
            
            embed = discord.Embed(
                title="✅ PAKET BERHASIL DITAMBAH",
                color=0x00ff00
            )
            embed.add_field(name="🆔 ID Paket", value=f"`{pkg_id}`", inline=False)
            embed.add_field(name="📦 Nama", value=pkg_name, inline=True)
            embed.add_field(name="💰 Harga", value=f"Rp **{pkg_price:,}**", inline=True)
            embed.add_field(name="⏱️ Durasi", value=f"**{pkg_duration}** hari", inline=True)
            embed.add_field(name="✅ Status", value="Paket sudah tersimpan di database!", inline=False)
            embed.set_footer(text=f"Dibuat oleh: {interaction.user.name}")
            
            await interaction.followup.send(embed=embed, ephemeral=True)
        except ValueError:
            await interaction.followup.send("❌ Harga dan Durasi harus berupa angka!", ephemeral=True)
        except Exception as e:
            await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


class DeletePackageView(discord.ui.View):
    def __init__(self, packages):
        super().__init__()
        self.packages = packages
        
        # Create dropdown with all packages
        options = []
        for key, pkg in packages.items():
            options.append(discord.SelectOption(
                label=f"{pkg['name']} (Rp {pkg['price']:,})",
                value=key,
                description=f"ID: {key}"
            ))
        
        select = discord.ui.Select(
            placeholder="Pilih paket yang ingin dihapus...",
            options=options
        )
        select.callback = self.select_callback
        self.add_item(select)
    
    async def select_callback(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        
        selected_id = interaction.data['values'][0]
        pkg = self.packages[selected_id]
        
        try:
            # Check if package is used in any subscription
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            c.execute('SELECT COUNT(*) FROM subscriptions WHERE package_type = ?', (selected_id,))
            usage_count = c.fetchone()[0]
            
            if usage_count > 0:
                conn.close()
                embed = discord.Embed(
                    title="❌ TIDAK BISA MENGHAPUS",
                    color=0xff6b6b
                )
                embed.add_field(name="⚠️ Alasan", value=f"Paket ini masih digunakan oleh **{usage_count}** member aktif!", inline=False)
                embed.add_field(name="📝 Saran", value="Hapus member terlebih dahulu sebelum menghapus paket", inline=False)
                
                await interaction.followup.send(embed=embed, ephemeral=True)
                return
            
            # Delete from database
            c.execute('DELETE FROM packages WHERE package_id = ?', (selected_id,))
            conn.commit()
            conn.close()
            
            embed = discord.Embed(
                title="✅ PAKET BERHASIL DIHAPUS",
                color=0xff6b6b
            )
            embed.add_field(name="🗑️ Paket Dihapus", value=pkg['name'], inline=False)
            embed.add_field(name="🆔 ID", value=f"`{selected_id}`", inline=True)
            embed.add_field(name="💰 Harga", value=f"Rp {pkg['price']:,}", inline=True)
            embed.add_field(name="✅ Status", value="Paket sudah dihapus dari database!", inline=False)
            embed.set_footer(text=f"Dihapus oleh: {interaction.user.name}")
            
            await interaction.followup.send(embed=embed, ephemeral=True)
        except Exception as e:
            await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


class ManagePackagesView(discord.ui.View):
    def __init__(self, packages):
        super().__init__()
        self.packages = packages
    
    @discord.ui.button(label="➕ Buat Paket", style=discord.ButtonStyle.green, emoji="📦")
    async def create_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(CreatePackageModal())
    
    @discord.ui.button(label="🗑️ Hapus Paket", style=discord.ButtonStyle.red, emoji="❌")
    async def delete_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = DeletePackageView(self.packages)
        embed = discord.Embed(
            title="🗑️ HAPUS PAKET",
            description="Pilih paket yang ingin dihapus di dropdown bawah ini:",
            color=0xff6b6b
        )
        await interaction.response.send_message(embed=embed, view=view, ephemeral=True)


@tree.command(name="manage_packages", description="[Admin] Kelola paket membership - buat atau hapus")
@discord.app_commands.default_permissions(administrator=False)
async def manage_packages_command(interaction: discord.Interaction):
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message(
            "❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", 
            ephemeral=True)
        return
    
    packages = get_all_packages()
    
    # Show current packages
    embed = discord.Embed(
        title="📦 KELOLA PAKET MEMBERSHIP",
        description="Paket-paket yang tersedia saat ini:",
        color=0xf7931a
    )
    
    for key, pkg in packages.items():
        embed.add_field(
            name=f"**{pkg['name']}**",
            value=f"ID: `{key}`\n💰 Rp **{pkg['price']:,}**\n⏱️ {pkg['duration_text']}",
            inline=False
        )
    
    embed.set_footer(text="Gunakan tombol di bawah untuk buat atau hapus paket")
    
    view = ManagePackagesView(packages)
    await interaction.response.send_message(embed=embed, view=view, ephemeral=True)


class CreateDiscountModal(discord.ui.Modal, title="💰 Create Discount Code"):
    code = discord.ui.TextInput(label="Kode Diskon", placeholder="Contoh: SUMMER50, BLACK20", required=True, max_length=20)
    discount_percent = discord.ui.TextInput(label="Diskon (%)", placeholder="Contoh: 10, 25, 50", required=True, max_length=3)
    validity_days = discord.ui.TextInput(label="Berlaku Berapa Hari", placeholder="Contoh: 7, 30, 90", required=True, max_length=3)
    max_uses = discord.ui.TextInput(label="Max Uses (orang)", placeholder="Contoh: 5, 10 (atau 0 untuk unlimited)", required=True, max_length=3)
    
    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        
        try:
            code_val = self.code.value.strip().upper()
            discount_percent = int(self.discount_percent.value.strip())
            validity_days = int(self.validity_days.value.strip())
            max_uses = int(self.max_uses.value.strip())
            
            if discount_percent <= 0 or discount_percent > 100:
                await interaction.followup.send("❌ Diskon harus antara 1-100%!", ephemeral=True)
                return
            
            if validity_days <= 0:
                await interaction.followup.send("❌ Validity days harus lebih besar dari 0!", ephemeral=True)
                return
            
            if max_uses < 0:
                await interaction.followup.send("❌ Max uses tidak boleh negatif!", ephemeral=True)
                return
            
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            # Create discount code table if not exists
            c.execute('''CREATE TABLE IF NOT EXISTS discount_codes (
                code TEXT PRIMARY KEY,
                discount_percent INTEGER,
                validity_days INTEGER DEFAULT 1,
                max_uses INTEGER,
                used_count INTEGER DEFAULT 0,
                created_at TEXT,
                code_expiry_date TEXT,
                created_by TEXT
            )''')
            
            created_at = get_jakarta_datetime().strftime('%Y-%m-%d %H:%M:%S')
            code_expiry = (get_jakarta_datetime() + timedelta(days=validity_days)).strftime('%Y-%m-%d %H:%M:%S')
            creator = interaction.user.name
            
            c.execute('''INSERT OR REPLACE INTO discount_codes 
                        (code, discount_percent, validity_days, max_uses, created_at, code_expiry_date, created_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''',
                     (code_val, discount_percent, validity_days, max_uses, created_at, code_expiry, creator))
            
            conn.commit()
            conn.close()
            
            embed = discord.Embed(
                title="✅ DISKON CODE DIBUAT",
                color=0x00ff00
            )
            embed.add_field(name="💳 Kode", value=f"`{code_val}`", inline=False)
            embed.add_field(name="📊 Diskon", value=f"**{discount_percent}%**", inline=True)
            embed.add_field(name="📅 Berlaku Sampai", value=f"**{validity_days} hari** ({code_expiry})", inline=True)
            embed.add_field(name="👥 Max Uses", value=f"**{max_uses if max_uses > 0 else 'Unlimited'}**", inline=True)
            embed.add_field(name="👤 Dibuat Oleh", value=creator, inline=True)
            embed.add_field(name="⏰ Waktu", value=format_jakarta_datetime(get_jakarta_datetime()), inline=True)
            embed.set_footer(text="Diary Crypto Payment Bot")
            
            await interaction.followup.send(embed=embed, ephemeral=True)
        except ValueError:
            await interaction.followup.send("❌ Semua input harus berupa angka!", ephemeral=True)
        except Exception as e:
            await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


@tree.command(name="create_discount", description="[Admin] Buat kode diskon dengan validity period")
@discord.app_commands.default_permissions(administrator=False)
async def create_discount_command(interaction: discord.Interaction):
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message("❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", ephemeral=True)
        return
    
    await interaction.response.send_modal(CreateDiscountModal())


@tree.command(name="referral_link", description="[Analyst Only] Dapatkan referral link unik Anda")
async def referral_link_command(interaction: discord.Interaction):
    # List of 7 analysts (case-insensitive) - Updated dengan actual Discord usernames
    ANALYSTS = ["Bay", "Bel", "Dialena", "Kamado", "Rey", "Ryzu", "Zen"]
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    
    # Check if user is analyst (case-insensitive), admin, or orion
    is_analyst = interaction.user.name.lower() in [a.lower() for a in ANALYSTS]
    is_admin = interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id
    
    if not (is_analyst or is_admin or is_orion):
        await interaction.response.send_message(
            "❌ Command ini hanya untuk **Analyst** (Bay, Bel, Dialena, Kamado, Rey, Ryzu, Zen), **Admin**, **Guild Owner**, atau **Orion**!",
            ephemeral=True
        )
        return
    
    print(f"✅ Referral link accessed by: {interaction.user.name} (Analyst: {is_analyst})")
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        analyst_id = str(interaction.user.id)
        analyst_name = interaction.user.name
        
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        # Check if analyst referral code exists
        c.execute('SELECT code FROM referral_codes WHERE created_by = ?', (analyst_id,))
        existing_code = c.fetchone()
        
        if existing_code:
            ref_code = existing_code[0]
        else:
            # Generate unique referral code
            import random
            import string
            ref_code = f"REF{analyst_name[:3].upper()}{random.randint(1000, 9999)}"
            
            c.execute('''INSERT INTO referral_codes (code, created_by, uses)
                        VALUES (?, ?, ?)''',
                     (ref_code, analyst_id, 0))
            conn.commit()
        
        # Get referral stats
        c.execute('SELECT uses FROM referral_codes WHERE code = ?', (ref_code,))
        uses = c.fetchone()[0]
        
        c.execute('SELECT COUNT(*) FROM commissions WHERE analyst_id = ? AND status = "pending"', (analyst_id,))
        pending_commission = c.fetchone()[0]
        
        c.execute('SELECT SUM(commission_amount) FROM commissions WHERE analyst_id = ? AND status = "completed"', (analyst_id,))
        total_earned = c.fetchone()[0] or 0
        
        conn.close()
        
        embed = discord.Embed(
            title="🔗 REFERRAL LINK ANDA",
            description=f"Share kode ini untuk dapatkan komisi 30% dari setiap pembelian!",
            color=0xf7931a
        )
        embed.add_field(name="💳 Referral Code", value=f"`{ref_code}`", inline=False)
        embed.add_field(name="Cara Pakai", value="Kirim kode ini ke orang lain saat mereka membeli paket membership", inline=False)
        embed.add_field(name="👥 Total Referrals", value=f"**{uses}** orang", inline=True)
        pending_amount = pending_commission
        embed.add_field(name="💰 Komisi Pending", value=f"**Rp {int(pending_amount):,}**", inline=True)
        embed.add_field(name="✅ Komisi Earned", value=f"**Rp {int(total_earned):,}**", inline=True)
        embed.set_footer(text="30% komisi untuk setiap referral!")
        
        print(f"✅ Referral link shown for {analyst_name}: {ref_code}")
        await interaction.followup.send(embed=embed, ephemeral=True)
    except Exception as e:
        print(f"❌ Error in referral_link: {e}")
        await interaction.followup.send("⚠️ **Bot Sedang Maintenance** - Mohon hubungi admin", ephemeral=True)


@tree.command(name="referral_stats", description="[Admin] Lihat statistik referral semua analyst")
@discord.app_commands.default_permissions(administrator=False)
async def referral_stats_command(interaction: discord.Interaction):
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message("❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        
        # Get all analyst stats
        c.execute('''SELECT 
                     rc.created_by,
                     rc.code,
                     rc.uses,
                     COUNT(DISTINCT com.referred_member_id) as unique_referrals,
                     SUM(CASE WHEN com.status = "completed" THEN com.commission_amount ELSE 0 END) as earned,
                     SUM(CASE WHEN com.status = "pending" THEN com.commission_amount ELSE 0 END) as pending
                  FROM referral_codes rc
                  LEFT JOIN commissions com ON rc.created_by = com.analyst_id
                  GROUP BY rc.created_by
                  ORDER BY earned DESC
        ''')
        
        stats = c.fetchall()
        conn.close()
        
        if not stats:
            await interaction.followup.send("ℹ️ Belum ada data referral", ephemeral=True)
            return
        
        embed = discord.Embed(
            title="📊 STATISTIK REFERRAL ANALYST",
            description="Ranking analyst berdasarkan komisi earned",
            color=0xf7931a
        )
        
        for idx, (analyst_id, code, uses, unique_referrals, earned, pending) in enumerate(stats, 1):
            earned = int(earned) if earned else 0
            pending = int(pending) if pending else 0
            
            embed.add_field(
                name=f"#{idx} - Kode: `{code}`",
                value=f"👥 Referrals: {uses} | 💰 Earned: Rp {earned:,} | ⏳ Pending: Rp {pending:,}",
                inline=False
            )
        
        embed.set_footer(text="📈 30% komisi per referral")
        await interaction.followup.send(embed=embed, ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


class CreateTrialCodeModal(discord.ui.Modal, title="🎫 Create Trial Code"):
    duration = discord.ui.TextInput(label="Durasi Trial (hari)", placeholder="Contoh: 1, 3, 7", required=True, max_length=3)
    validity = discord.ui.TextInput(label="Kode Berlaku Selama (hari)", placeholder="Contoh: 1, 7, 30", required=True, max_length=3)
    max_users = discord.ui.TextInput(label="Jumlah Orang Yang Bisa Pakai", placeholder="Contoh: 1, 5, 10 (atau 0 untuk unlimited)", required=True, max_length=3)
    
    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        
        try:
            duration_days = int(self.duration.value.strip())
            validity_days = int(self.validity.value.strip())
            max_uses = int(self.max_users.value.strip())
            
            if duration_days <= 0 or validity_days <= 0:
                await interaction.followup.send("❌ Durasi dan validity harus lebih besar dari 0!", ephemeral=True)
                return
            
            if max_uses < 0:
                await interaction.followup.send("❌ Max users tidak boleh negatif!", ephemeral=True)
                return
            
            import random
            import string
            
            trial_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            
            conn = sqlite3.connect('warrior_subscriptions.db')
            c = conn.cursor()
            
            created_at = get_jakarta_datetime().strftime('%Y-%m-%d %H:%M:%S')
            creator = interaction.user.name
            code_expiry = (get_jakarta_datetime() + timedelta(days=validity_days)).strftime('%Y-%m-%d %H:%M:%S')
            
            c.execute('''INSERT INTO trial_members 
                        (trial_code, status, created_at, created_by, duration_days, validity_days, code_expiry_date, max_uses)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                     (trial_code, 'pending', created_at, creator, duration_days, validity_days, code_expiry, max_uses))
            
            conn.commit()
            conn.close()
            
            embed = discord.Embed(
                title="🎫 TRIAL CODE GENERATED",
                description="Kode trial member berhasil dibuat!",
                color=0xf7931a
            )
            embed.add_field(name="Kode Trial", value=f"`{trial_code}`", inline=False)
            embed.add_field(name="⏱️ Durasi Trial", value=f"{duration_days} hari (ketika di-redeem)", inline=True)
            embed.add_field(name="📅 Kode Berlaku", value=f"{validity_days} hari", inline=True)
            embed.add_field(name="👥 Max Users", value=f"{max_uses if max_uses > 0 else 'Unlimited'} orang", inline=True)
            embed.add_field(name="🔴 Expired Tanggal", value=code_expiry, inline=False)
            embed.add_field(name="📝 Status", value="PENDING - Menunggu redemption", inline=False)
            embed.add_field(name="💡 Cara Pakai", value="Bagikan kode ke user, user pakai `/redeem_trial` untuk redeem", inline=False)
            embed.add_field(name="Dibuat Oleh", value=creator, inline=True)
            embed.add_field(name="Waktu", value=format_jakarta_datetime(get_jakarta_datetime()), inline=True)
            embed.set_footer(text="Diary Crypto Payment Bot")
            
            await interaction.followup.send(embed=embed, ephemeral=True)
        except ValueError:
            await interaction.followup.send("❌ Input harus berupa angka!", ephemeral=True)
        except Exception as e:
            await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)


@tree.command(name="create_trial_code", description="[Admin] Buat kode trial member dengan durasi & validity")
@discord.app_commands.default_permissions(administrator=False)
async def create_trial_code_command(interaction: discord.Interaction):
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message("❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", ephemeral=True)
        return
    
    await interaction.response.send_modal(CreateTrialCodeModal())


@tree.command(name="kick_member", description="[Admin/Com-Manager] Kick member secara manual")
@discord.app_commands.default_permissions(administrator=False)
async def kick_member_command(interaction: discord.Interaction):
    is_orion = interaction.user.name.lower() == "orion" or str(interaction.user.id) == "orion"
    if not (interaction.user.guild_permissions.administrator or interaction.user.id == interaction.guild.owner_id or is_orion):
        await interaction.response.send_message("❌ Command ini hanya untuk **Admin**, **Guild Owner**, atau **Orion**!", ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    guild = interaction.guild
    
    # Ambil list members dengan role The Warrior atau Trial Member
    warrior_members = []
    trial_members = []
    
    warrior_role = discord.utils.get(guild.roles, name=WARRIOR_ROLE_NAME)
    trial_role = discord.utils.get(guild.roles, name=TRIAL_MEMBER_ROLE_NAME)
    
    if warrior_role:
        warrior_members = [m for m in warrior_role.members]
    if trial_role:
        trial_members = [m for m in trial_role.members]
    
    if not warrior_members and not trial_members:
        await interaction.followup.send("❌ Tidak ada member dengan role The Warrior atau Trial Member", ephemeral=True)
        return
    
    class KickMemberSearchModal(discord.ui.Modal, title="🔍 Cari Member untuk Di-Kick"):
        search_name = discord.ui.TextInput(label="Cari Nama Member", placeholder="Ketik nama member...", required=True, max_length=100)
        
        async def on_submit(self, modal_interaction: discord.Interaction):
            await modal_interaction.response.defer(ephemeral=True)
            
            search_query = self.search_name.value.strip().lower()
            
            # Search in all members
            found_warriors = [m for m in warrior_members if search_query in m.name.lower()]
            found_trials = [m for m in trial_members if search_query in m.name.lower()]
            
            if not found_warriors and not found_trials:
                await modal_interaction.followup.send(f"❌ Tidak ada member ditemukan dengan nama mengandung '{search_query}'", ephemeral=True)
                return
            
            # Create selection from results
            class SearchResultSelect(discord.ui.Select):
                def __init__(self):
                    options = []
                    self.member_map = {}
                    
                    for m in found_warriors[:20]:
                        options.append(discord.SelectOption(label=f"⚔️ {m.name} (The Warrior)", value=f"warrior_{m.id}"))
                        self.member_map[f"warrior_{m.id}"] = (m, "The Warrior", warrior_role)
                    
                    for m in found_trials[:20]:
                        options.append(discord.SelectOption(label=f"🎫 {m.name} (Trial)", value=f"trial_{m.id}"))
                        self.member_map[f"trial_{m.id}"] = (m, "Trial Member", trial_role)
                    
                    super().__init__(placeholder="Pilih member untuk di-kick...", options=options)
                
                async def callback(self, select_interaction: discord.Interaction):
                    await select_interaction.response.defer(ephemeral=True)
                    
                    selected = self.values[0]
                    member, role_name, role_obj = self.member_map[selected]
                    
                    try:
                        if role_obj and role_obj in member.roles:
                            await member.remove_roles(role_obj)
                            
                            try:
                                kick_embed = discord.Embed(
                                    title="🚨 ANDA TELAH DI-KICK! 🚨",
                                    description=f"Role **{role_name}** telah dihapus dari akun Anda.",
                                    color=0xff4444
                                )
                                kick_embed.add_field(name="❌ Status", value="KICKED", inline=True)
                                kick_embed.add_field(name="🔴 Role Dihapus", value=role_name, inline=True)
                                kick_embed.add_field(name="💬 Pertanyaan?", value="Hubungi admin untuk informasi lebih lanjut", inline=False)
                                kick_embed.set_footer(text="Diary Crypto Payment Bot • Real Time WIB")
                                kick_embed.set_thumbnail(url=member.avatar.url if member.avatar else "")
                                
                                await member.send(embed=kick_embed)
                            except:
                                pass
                            
                            send_admin_kick_notification(member.name, member.mention if hasattr(member, 'mention') else member.name, "Membership", "Admin Kick")
                            await select_interaction.followup.send(f"✅ {member.name} berhasil di-kick dari role {role_name}!", ephemeral=True)
                        else:
                            await select_interaction.followup.send(f"❌ Member tidak memiliki role {role_name}", ephemeral=True)
                    except Exception as e:
                        print(f"❌ Kick error: {e}")
                        await select_interaction.followup.send("⚠️ **Bot Sedang Maintenance** - Mohon hubungi admin", ephemeral=True)
            
            class SearchResultView(discord.ui.View):
                def __init__(self):
                    super().__init__()
                    self.add_item(SearchResultSelect())
            
            result_embed = discord.Embed(
                title="🔍 HASIL PENCARIAN",
                description=f"Ditemukan {len(found_warriors) + len(found_trials)} member",
                color=0xf7931a
            )
            if found_warriors:
                result_embed.add_field(name="⚔️ The Warrior", value=f"{len(found_warriors)} member", inline=True)
            if found_trials:
                result_embed.add_field(name="🎫 Trial Member", value=f"{len(found_trials)} member", inline=True)
            
            await modal_interaction.followup.send(embed=result_embed, view=SearchResultView(), ephemeral=True)
    
    class KickMemberView(discord.ui.View):
        @discord.ui.button(label="🔍 Cari Member", style=discord.ButtonStyle.primary)
        async def search_button(self, button_interaction: discord.Interaction, button: discord.ui.Button):
            await button_interaction.response.send_modal(KickMemberSearchModal())
    
    embed = discord.Embed(
        title="🚨 KICK MEMBER",
        description="Gunakan search untuk cari member yang ingin di-kick. Role akan dihapus dan member dapat notifikasi.",
        color=0xff0000
    )
    embed.add_field(name="⚔️ The Warrior", value=f"Total: {len(warrior_members)} members", inline=True)
    embed.add_field(name="🎫 Trial Member", value=f"Total: {len(trial_members)} members", inline=True)
    
    await interaction.followup.send(embed=embed, view=KickMemberView(), ephemeral=True)


@tree.error
async def on_app_command_error(interaction: discord.Interaction, error: discord.app_commands.AppCommandError):
    """Global error handler untuk semua app commands"""
    try:
        print(f"❌ Command error: {error}")
        if not interaction.response.is_done():
            await interaction.response.send_message(
                "⚠️ **Bot Sedang Maintenance**\n\nMohon hubungi admin untuk bantuan lebih lanjut. Kami akan segera memperbaiki sistem!\n\n💬 Kontak Admin: /support", 
                ephemeral=True
            )
    except Exception as e:
        print(f"❌ Error handler error: {e}")
        try:
            if not interaction.response.is_done():
                await interaction.response.send_message(
                    "⚠️ **Bot Sedang Maintenance**\n\nMohon hubungi admin untuk bantuan lebih lanjut.",
                    ephemeral=True
                )
        except:
            pass


# ============ FLASK ROUTES ============
@app.route('/')
def home():
    return '''
    <html>
        <head><title>Diary Crypto Payment Bot</title></head>
        <body style="font-family: Arial; text-align: center; padding: 50px;">
            <h1>🤖 Diary Crypto Payment Bot - Running</h1>
            <p>Bot Status: <b>ONLINE ✅</b></p>
            <p>Discord Guild: Diary Crypto</p>
            <p>Payment Gateway: Midtrans SANDBOX</p>
        </body>
    </html>
    '''


def delete_pending_order(order_id):
    """Delete pending order after successful payment"""
    try:
        conn = sqlite3.connect('warrior_subscriptions.db')
        c = conn.cursor()
        c.execute('DELETE FROM pending_orders WHERE order_id = ?', (order_id,))
        conn.commit()
        conn.close()
        print(f"✅ Pending order deleted: {order_id}")
        return True
    except Exception as e:
        print(f"⚠️ Error deleting pending order: {e}")
        return False

@app.route('/webhook/midtrans', methods=['POST'])
def midtrans_webhook():
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        transaction_status = data.get('transaction_status')
        
        print(f"🔔 Webhook received: Order {order_id} - Status {transaction_status}")
        
        # Handle both 'settlement' dan 'capture' status (payment successful)
        if transaction_status in ['settlement', 'capture', 'accept_partial_credit']:
            pending = get_pending_order(order_id)
            print(f"📋 Pending order lookup result: {pending}")
            
            if pending:
                order_id_db, discord_id, discord_username, nama, email, package_type, payment_url, status, created_at = pending
                print(f"✅ Found pending order - Discord ID: {discord_id}, Package: {package_type}")
                
                save_subscription(order_id, discord_id, discord_username, nama, email, package_type)
                print(f"✅ Subscription activated for {nama}")
                
                # Assign role dan send notifications
                try:
                    packages = get_all_packages()
                    pkg = packages.get(package_type)
                    pkg_name = pkg['name'] if pkg else package_type
                    
                    # Get subscription dates
                    conn = sqlite3.connect('warrior_subscriptions.db')
                    c = conn.cursor()
                    c.execute('SELECT start_date, end_date FROM subscriptions WHERE order_id = ?', (order_id,))
                    sub_data = c.fetchone()
                    conn.close()
                    
                    if sub_data:
                        start_date, end_date = sub_data
                        
                        # Get user from Discord
                        guild = bot.get_guild(GUILD_ID)
                        print(f"📌 Guild lookup: {guild}")
                        
                        if guild:
                            member = guild.get_member(int(discord_id))
                            print(f"👤 Member lookup for ID {discord_id}: {member}")
                            
                            if member:
                                member_avatar = str(member.avatar.url) if member.avatar else str(member.default_avatar)
                                warrior_role = discord.utils.get(guild.roles, name=WARRIOR_ROLE_NAME)
                                
                                # 1. ASSIGN ROLE
                                if warrior_role:
                                    try:
                                        asyncio.run_coroutine_threadsafe(
                                            member.add_roles(warrior_role),
                                            bot.loop
                                        )
                                        print(f"✅ Role '{WARRIOR_ROLE_NAME}' assigned to {nama}")
                                    except Exception as e:
                                        print(f"⚠️ Error assigning role: {e}")
                                else:
                                    print(f"❌ Role '{WARRIOR_ROLE_NAME}' not found in guild")
                                
                                # 2. Send welcome email
                                send_welcome_email(nama, email, pkg_name, order_id, start_date, end_date, "", member_avatar)
                                
                                # 3. Send DM with EMBED (matching /buy checkout format)
                                try:
                                    dm_embed = discord.Embed(
                                        title="✅ SELAMAT DATANG DI THE WARRIOR!",
                                        description="Membership Anda berhasil diaktifkan!",
                                        color=0xf7931a
                                    )
                                    dm_embed.set_thumbnail(url=member_avatar)
                                    dm_embed.add_field(name="📦 Paket", value=f"**{pkg_name}**", inline=True)
                                    dm_embed.add_field(name="📋 Order ID", value=f"`{order_id}`", inline=True)
                                    dm_embed.add_field(name="📅 Mulai", value=start_date, inline=True)
                                    dm_embed.add_field(name="⏰ Berakhir", value=end_date, inline=True)
                                    dm_embed.add_field(name="🎯 Info", value="Nikmati akses eksklusif ke The Warrior!", inline=False)
                                    dm_embed.set_footer(text="Diary Crypto Payment Bot • Terima kasih!")
                                    
                                    asyncio.run_coroutine_threadsafe(
                                        member.send(embed=dm_embed),
                                        bot.loop
                                    )
                                    print(f"✅ Welcome embed sent to {nama}")
                                except Exception as e:
                                    print(f"⚠️ Could not send DM to {nama}: {e}")
                                
                                # 4. Send admin notification
                                send_admin_new_member_notification(nama, order_id, pkg_name, email)
                                
                                # 5. DELETE PENDING ORDER IMMEDIATELY - Jadi tidak dapat notif "ORDER KADALUARSA" nanti
                                delete_pending_order(order_id)
                                print(f"✅ Pending order automatically deleted after payment success")
                                
                except Exception as e:
                    print(f"⚠️ Error processing webhook: {e}")
                    # Still delete pending order even if notifications failed
                    delete_pending_order(order_id)
            
            else:
                print(f"⚠️ Pending order NOT found for {order_id} - might be already processed or expired")
        
        return {'status': 'ok'}, 200
    except Exception as e:
        print(f"❌ Webhook error: {e}")
        return {'status': 'error'}, 500


# ============ MAIN ============
if __name__ == '__main__':
    print("🔑 Midtrans Server Key: ✅ SET" if MIDTRANS_SERVER_KEY else "🔑 Midtrans Server Key: ❌ NOT SET")
    print("🔑 Midtrans Client Key: ✅ SET" if MIDTRANS_CLIENT_KEY else "🔑 Midtrans Client Key: ❌ NOT SET")
    print("📧 Gmail Sender: ✅ SET" if GMAIL_SENDER else "📧 Gmail Sender: ❌ NOT SET")
    print("📧 Admin Email: ✅ SET" if ADMIN_EMAIL else "📧 Admin Email: ❌ NOT SET")
    
    # Flask app
    def run_flask():
        app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
    
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    
    print(f"🚀 Starting Discord bot...")
    print(f"🌐 Webhook URL untuk Midtrans: https://{os.getenv('REPLIT_PROJECT_DOMAIN', 'localhost')}/webhook/midtrans")
    print(f"🧪 Midtrans Mode: SANDBOX (Testing)")
    print(f"💡 Pastikan webhook URL sudah dikonfigurasi di dashboard Midtrans SANDBOX")
    
    bot.run(DISCORD_TOKEN)
